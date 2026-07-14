"""
Kouzina Live Updates — admin activity feed + private 1:1 direct messages.

Auth:
    Admins log in with their pkdb.admins email + password (bcrypt). On success
    they get a signed JWT (HS256, signed with APP_SECRET) used as a Bearer token.

Live Updates:
    A Twitter-style feed of actions performed on the portals. Events are pulled
    from existing audit tables (currently pkdb.admin_audit_log) into a
    normalized feed_events table. Ingestion happens on-read: when an admin opens
    the feed and the last ingest is older than INGEST_INTERVAL, the app pulls
    any new audit rows right then (guarded by a MySQL advisory lock so two
    replicas never double-ingest). Admins can like and comment on events.

Messaging:
    Messages are private between two admins. You only ever see conversations you
    are part of, and you can only message active admins (no external parties).

Storage:
    App data lives in the app's database (DATABASE_URL / MYSQL_*). Admins and
    audit logs are read from pkdb on the SAME MySQL server via cross-database
    queries, so no separate connection is needed.
"""

import io
import json
import os
import smtplib
import ssl
from datetime import datetime, timedelta, timezone
from email.message import EmailMessage
from pathlib import Path
from typing import Optional

import bcrypt
import jwt
from fastapi import (
    BackgroundTasks, Depends, FastAPI, File, HTTPException, Query, UploadFile,
)
from fastapi.middleware.cors import CORSMiddleware
from fastapi.security import HTTPAuthorizationCredentials, HTTPBearer
from fastapi.staticfiles import StaticFiles
from pydantic import BaseModel, Field
from sqlalchemy import (
    URL,
    Column,
    DateTime,
    Integer,
    String,
    Text,
    UniqueConstraint,
    and_,
    bindparam,
    create_engine,
    func,
    or_,
    text,
)
from sqlalchemy.orm import declarative_base, sessionmaker

# ---- Database --------------------------------------------------------------


def _build_db_url():
    host = os.environ.get("MYSQL_HOST")
    if host:
        return URL.create(
            "mysql+pymysql",
            username=os.environ.get("MYSQL_USER"),
            password=os.environ.get("MYSQL_PASSWORD"),
            host=host,
            port=int(os.environ.get("MYSQL_PORT", "3306")),
            database=os.environ.get("MYSQL_DB"),
        )
    env_url = os.environ.get("DATABASE_URL")
    if env_url:
        return env_url
    return f"sqlite:///{Path(__file__).parent / 'messages.db'}"


DATABASE_URL = _build_db_url()
_backend = (
    DATABASE_URL.drivername
    if isinstance(DATABASE_URL, URL)
    else str(DATABASE_URL).split("://", 1)[0]
)
IS_MYSQL = _backend.startswith("mysql")

connect_args = {}
if IS_MYSQL:
    ca = os.environ.get("MYSQL_SSL_CA")
    connect_args = {"ssl": {"ca": ca} if ca else {"ssl": True}}
elif _backend.startswith("sqlite"):
    connect_args = {"check_same_thread": False}

engine = create_engine(DATABASE_URL, pool_pre_ping=True, connect_args=connect_args)
SessionLocal = sessionmaker(bind=engine, autoflush=False, expire_on_commit=False)
Base = declarative_base()


class MessageRow(Base):
    __tablename__ = "messages"
    id = Column(Integer, primary_key=True, autoincrement=True)
    sender = Column(String(255), nullable=False)
    recipient = Column(String(255), nullable=False)
    body = Column(Text, nullable=False)
    is_private = Column(Integer, nullable=False, default=0)  # 1 = sender+recipient only
    parent_id = Column(Integer, nullable=True, index=True)   # reply threading
    created_at = Column(DateTime(timezone=True), server_default=func.now())


class FeedEventRow(Base):
    __tablename__ = "feed_events"
    __table_args__ = (UniqueConstraint("source", "source_id", name="uq_source_row"),)
    id = Column(Integer, primary_key=True, autoincrement=True)
    portal = Column(String(16), nullable=False)          # 'PK', 'KAC', 'KFC', ...
    source = Column(String(64), nullable=False)          # e.g. 'pkdb.admin_audit_log'
    source_id = Column(Integer, nullable=False)
    actor = Column(String(255), nullable=False)
    action = Column(String(64), nullable=False)
    summary = Column(String(512), nullable=False)
    details = Column(Text, nullable=True)
    happened_at = Column(DateTime(timezone=True), nullable=True, index=True)


class EventLikeRow(Base):
    __tablename__ = "event_likes"
    __table_args__ = (UniqueConstraint("event_id", "admin_email", name="uq_event_admin"),)
    id = Column(Integer, primary_key=True, autoincrement=True)
    event_id = Column(Integer, nullable=False, index=True)
    admin_email = Column(String(255), nullable=False)
    created_at = Column(DateTime(timezone=True), server_default=func.now())


class EventCommentRow(Base):
    __tablename__ = "event_comments"
    id = Column(Integer, primary_key=True, autoincrement=True)
    event_id = Column(Integer, nullable=False, index=True)
    admin_email = Column(String(255), nullable=False)
    admin_name = Column(String(255), nullable=True)
    body = Column(Text, nullable=False)
    created_at = Column(DateTime(timezone=True), server_default=func.now())


class ProgramRow(Base):
    __tablename__ = "programs"
    id = Column(Integer, primary_key=True, autoincrement=True)
    name = Column(String(255), nullable=False)
    objective = Column(Text, nullable=True)
    description = Column(Text, nullable=True)
    owner_email = Column(String(255), nullable=True)
    owner_name = Column(String(255), nullable=True)
    department = Column(String(255), nullable=True)
    eta = Column(DateTime(timezone=True), nullable=True)
    status = Column(String(20), nullable=False, default="not_started")
    active = Column(Integer, nullable=False, default=1)
    created_by = Column(String(255), nullable=True)
    created_at = Column(DateTime(timezone=True), server_default=func.now())


class ProgramUpdateRow(Base):
    __tablename__ = "program_updates"
    id = Column(Integer, primary_key=True, autoincrement=True)
    program_id = Column(Integer, nullable=False, index=True)
    author_email = Column(String(255), nullable=False)
    author_name = Column(String(255), nullable=True)
    body = Column(Text, nullable=False)
    created_at = Column(DateTime(timezone=True), server_default=func.now())


class FeedbackRow(Base):
    __tablename__ = "feedback"
    # Deliberately NO author column — feedback is anonymous by design.
    id = Column(Integer, primary_key=True, autoincrement=True)
    body = Column(Text, nullable=False)
    action_item = Column(Text, nullable=True)
    action_by = Column(String(255), nullable=True)
    action_at = Column(DateTime(timezone=True), nullable=True)
    status = Column(String(16), nullable=False, default="open")
    created_at = Column(DateTime(timezone=True), server_default=func.now())


class LiveLoginRow(Base):
    __tablename__ = "live_logins"
    id = Column(Integer, primary_key=True, autoincrement=True)
    email = Column(String(255), nullable=False, index=True)
    name = Column(String(255), nullable=True)
    created_at = Column(DateTime(timezone=True), server_default=func.now())


class MessageLikeRow(Base):
    __tablename__ = "message_likes"
    __table_args__ = (UniqueConstraint("message_id", "admin_email", name="uq_msg_admin"),)
    id = Column(Integer, primary_key=True, autoincrement=True)
    message_id = Column(Integer, nullable=False, index=True)
    admin_email = Column(String(255), nullable=False)
    created_at = Column(DateTime(timezone=True), server_default=func.now())


class RedflagTemplateRow(Base):
    __tablename__ = "redflag_templates"
    rule_key = Column(String(64), primary_key=True)
    subject = Column(Text, nullable=True)
    body = Column(Text, nullable=True)
    updated_by = Column(String(255), nullable=True)
    updated_at = Column(DateTime(timezone=True), server_default=func.now())


class IngestStateRow(Base):
    __tablename__ = "ingest_state"
    source = Column(String(64), primary_key=True)
    last_id = Column(Integer, nullable=False, default=0)
    last_run = Column(DateTime(timezone=True), nullable=True)


class UserProfileRow(Base):
    """KLU-side org profile for a pkdb admin. The account itself (email,
    password, active) lives in pkdb.admins — permissions stay KPK-managed."""
    __tablename__ = "user_profiles"
    id = Column(Integer, primary_key=True, autoincrement=True)
    email = Column(String(255), nullable=False, unique=True, index=True)
    function = Column(String(255), nullable=True)
    department = Column(String(255), nullable=True)
    sub_department = Column(String(255), nullable=True)
    owner = Column(String(255), nullable=True)       # department owner (person)
    notes = Column(Text, nullable=True)
    created_by = Column(String(255), nullable=True)
    updated_at = Column(DateTime(timezone=True), server_default=func.now())


class OrgDeptRow(Base):
    """Company org structure (Function → Department → Leader → Owner),
    editable by superadmins; seeds the dropdowns and the feed filter."""
    __tablename__ = "org_departments"
    id = Column(Integer, primary_key=True, autoincrement=True)
    function = Column(String(255), nullable=False)
    department = Column(String(255), nullable=False)
    leader = Column(String(255), nullable=True)
    owner = Column(String(255), nullable=True)
    active = Column(Integer, nullable=False, default=1)
    updated_at = Column(DateTime(timezone=True), server_default=func.now())


Base.metadata.create_all(engine)

# Existing deployments created sender/recipient as VARCHAR(50); emails need more.
if IS_MYSQL:
    with engine.begin() as conn:
        for col in ("sender", "recipient"):
            try:
                conn.execute(text(f"ALTER TABLE messages MODIFY {col} VARCHAR(255) NOT NULL"))
            except Exception:
                pass  # already widened
        for ddl in (
            "ALTER TABLE messages ADD COLUMN is_private TINYINT(1) NOT NULL DEFAULT 0",
            "ALTER TABLE messages ADD COLUMN parent_id INT NULL",
            "ALTER TABLE programs ADD COLUMN objective TEXT NULL",
            "ALTER TABLE programs ADD COLUMN description TEXT NULL",
            "ALTER TABLE programs ADD COLUMN department VARCHAR(255) NULL",
        ):
            try:
                conn.execute(text(ddl))
            except Exception:
                pass  # already added

# Backfill: pre-existing programs inherit their owner's department (idempotent).
try:
    with SessionLocal() as _db:
        _orphans = (
            _db.query(ProgramRow)
            .filter(ProgramRow.department.is_(None), ProgramRow.owner_email.isnot(None))
            .all()
        )
        if _orphans:
            _pmap = {
                (p.email or "").lower(): p.department
                for p in _db.query(UserProfileRow).all()
                if p.department
            }
            _dirty = False
            for _r in _orphans:
                _d = _pmap.get((_r.owner_email or "").lower())
                if _d:
                    _r.department = _d
                    _dirty = True
            if _dirty:
                _db.commit()
except Exception:
    pass


def _iso_utc(dt) -> str:
    """Serialize a DB datetime as ISO-8601 with an explicit UTC offset. The
    databases store naive UTC; without the offset, browsers parse the string
    as LOCAL time and every timestamp shifts by the viewer's UTC offset."""
    if not dt:
        return ""
    if dt.tzinfo is None:
        dt = dt.replace(tzinfo=timezone.utc)
    return dt.isoformat()


# ---- Auth helpers ----------------------------------------------------------

APP_SECRET = os.environ.get("APP_SECRET", "dev-only-insecure-secret-change-me")
JWT_ALG = "HS256"
TOKEN_TTL = timedelta(hours=12)
SUPERADMIN_EMAIL = os.environ.get("SUPERADMIN_EMAIL", "admin@kftpl.com")

# System/bot accounts hidden from people lists and not messageable.
EXCLUDED_ADMIN_EMAILS = {"cocoadmin@kftpl.com", "swiggy-review@kftpl.com"}


def _excluded(email: Optional[str]) -> bool:
    return (email or "").lower() in EXCLUDED_ADMIN_EMAILS


def _smtp_config() -> Optional[dict]:
    """SMTP settings, preferring explicit env vars, then reusing PartnerKart's
    own mailer from pkdb.smtp_settings (same MySQL server). Reading it live
    means no credentials are copied anywhere and PK key rotations are picked up
    automatically."""
    host = os.environ.get("SMTP_HOST")
    if host:
        return {
            "host": host,
            "port": int(os.environ.get("SMTP_PORT", "587")),
            "user": os.environ.get("SMTP_USER"),
            "password": os.environ.get("SMTP_PASS"),
            "from_email": os.environ.get("SMTP_FROM") or os.environ.get("SMTP_USER"),
            "from_name": os.environ.get("SMTP_FROM_NAME", "Kouzina Live"),
        }
    if IS_MYSQL:
        try:
            with engine.connect() as conn:
                r = conn.execute(
                    text(
                        "SELECT smtp_host, smtp_port, smtp_username, smtp_password, "
                        "from_email, from_name FROM pkdb.smtp_settings "
                        "WHERE smtp_host IS NOT NULL ORDER BY id DESC LIMIT 1"
                    )
                ).mappings().first()
            if r and r["smtp_host"]:
                return {
                    "host": r["smtp_host"],
                    "port": r["smtp_port"] or 587,
                    "user": r["smtp_username"],
                    "password": r["smtp_password"],
                    "from_email": r["from_email"] or r["smtp_username"],
                    # Send under the Kouzina Live name but PK's verified address.
                    "from_name": "Kouzina Live",
                }
        except Exception:
            pass
    return None


def email_enabled() -> bool:
    return _smtp_config() is not None


def send_email(recipients: list[str], subject: str, body: str) -> bool:
    """Best-effort email via the resolved SMTP config. No-op if none available."""
    cfg = _smtp_config()
    if not cfg or not recipients:
        return False
    try:
        msg = EmailMessage()
        msg["From"] = f'{cfg["from_name"]} <{cfg["from_email"]}>'
        msg["To"] = ", ".join(recipients)
        msg["Subject"] = subject
        msg.set_content(body)
        ctx = ssl.create_default_context()
        with smtplib.SMTP(cfg["host"], cfg["port"], timeout=20) as s:
            s.starttls(context=ctx)
            if cfg["user"]:
                s.login(cfg["user"], cfg["password"])
            s.send_message(msg)
        return True
    except Exception:
        return False


def resolve_names(emails: set[str]) -> dict[str, str]:
    """email -> display name, for the emails that belong to admins."""
    if not emails:
        return {}
    if IS_MYSQL:
        with engine.connect() as conn:
            pairs = conn.execute(
                text("SELECT email, name FROM pkdb.admins WHERE email IN :emails")
                .bindparams(bindparam("emails", expanding=True)),
                {"emails": sorted(emails)},
            ).all()
        return {em: n for em, n in pairs if n}
    return {a["email"]: a["name"] for a in list_active_admins()}


def _verify_password(password: str, password_hash: str) -> bool:
    # Laravel/PHP often emit "$2y$" bcrypt hashes; normalise to "$2b$".
    if password_hash.startswith("$2y$"):
        password_hash = "$2b$" + password_hash[4:]
    try:
        return bcrypt.checkpw(password.encode(), password_hash.encode())
    except Exception:
        return False


def fetch_admin_by_email(email: str) -> Optional[dict]:
    if IS_MYSQL:
        with engine.connect() as conn:
            row = conn.execute(
                text(
                    "SELECT name, email, password_hash, active "
                    "FROM pkdb.admins WHERE email = :e LIMIT 1"
                ),
                {"e": email},
            ).mappings().first()
            return dict(row) if row else None
    # Local dev fallback (no pkdb available).
    dev_email = os.environ.get("DEV_ADMIN_EMAIL", "dev@local")
    if email.lower() == dev_email.lower():
        pw = os.environ.get("DEV_ADMIN_PASSWORD", "dev")
        return {
            "name": "Dev Admin",
            "email": dev_email,
            "password_hash": bcrypt.hashpw(pw.encode(), bcrypt.gensalt()).decode(),
            "active": 1,
        }
    return None


def list_active_admins() -> list[dict]:
    if IS_MYSQL:
        with engine.connect() as conn:
            rows = conn.execute(
                text("SELECT name, email FROM pkdb.admins WHERE active = 1 ORDER BY name")
            ).mappings().all()
            return [dict(r) for r in rows if not _excluded(r["email"])]
    return [
        {"name": "Dev Admin", "email": os.environ.get("DEV_ADMIN_EMAIL", "dev@local")},
        {"name": "Alice Admin", "email": "alice@local"},
    ]


def is_active_admin(email: str) -> bool:
    if _excluded(email):
        return False
    if IS_MYSQL:
        with engine.connect() as conn:
            return (
                conn.execute(
                    text("SELECT 1 FROM pkdb.admins WHERE email = :e AND active = 1 LIMIT 1"),
                    {"e": email},
                ).first()
                is not None
            )
    return any(a["email"] == email for a in list_active_admins())


def make_token(email: str, name: Optional[str]) -> str:
    payload = {
        "sub": email,
        "name": name or email,
        "exp": datetime.now(timezone.utc) + TOKEN_TTL,
    }
    return jwt.encode(payload, APP_SECRET, algorithm=JWT_ALG)


bearer = HTTPBearer(auto_error=False)


def current_admin(cred: Optional[HTTPAuthorizationCredentials] = Depends(bearer)) -> dict:
    if cred is None:
        raise HTTPException(status_code=401, detail="Not authenticated")
    try:
        payload = jwt.decode(cred.credentials, APP_SECRET, algorithms=[JWT_ALG])
    except jwt.PyJWTError:
        raise HTTPException(status_code=401, detail="Invalid or expired session")
    return {"email": payload["sub"], "name": payload.get("name")}


# ---- Feed ingestion ---------------------------------------------------------

INGEST_INTERVAL = timedelta(seconds=30)
INGEST_BATCH = 500
# Same person doing the same action within this window collapses to one card.
GROUP_WINDOW = timedelta(minutes=15)

# Routine events that would just be noise in the feed.
EXCLUDED_ACTIONS = {"login", "logout", "switch_tenant"}

# Friendlier phrasing for known actions; anything else falls back to
# "did <action with spaces>".
ACTION_LABELS = {
    "item_setting_changed": "changed settings on",
    "create": "created",
    "update": "updated",
    "delete": "deleted",
    "user_create": "created",
    "user_disable": "disabled",
}


def _summarize(action: str, entity_type: Optional[str], entity_id) -> str:
    verb = ACTION_LABELS.get(action, action.replace("_", " "))
    target = ""
    if entity_type:
        target = f" {entity_type}"
        if entity_id:
            target += f" #{entity_id}"
    return f"{verb}{target}".strip()


def _render_summary(action: str, summary: str, details: Optional[str]) -> str:
    """Enrich the stored summary from the audit row's details JSON at serve
    time — names beat ids, and before→after changes are what admins care about."""
    if not details:
        return summary
    try:
        d = json.loads(details)
    except (ValueError, TypeError):
        return summary
    if not isinstance(d, dict):
        return summary

    name = d.get("item_name") or d.get("name")
    module = d.get("module")  # which portal tab/section the action happened in

    # An item being switched on/off is a stock-availability signal — call it out.
    active = d.get("active")
    if action == "item_setting_changed" and isinstance(active, dict) and "to" in active:
        state = "ON" if active["to"] else "OFF"
        out = f"switched {state} · {name}" if name else f"switched {state} an item"
        return f"{out} · in {module}" if module else out

    changes = [
        f"{k}: {v['from']} → {v['to']}"
        for k, v in d.items()
        if isinstance(v, dict) and "from" in v and "to" in v and k != "module"
    ]
    out = summary
    if name:
        out = f"{out} · {name}"
    if changes:
        out = f"{out} ({', '.join(changes[:2])})"
    if module:
        out = f"{out} · in {module}"
    return out


def _init_sql(table: str, id_col: str = "id", time_col: str = "created_at") -> str:
    """Watermark for a source's first-ever ingest: backfill at most the newest
    100 rows AND nothing older than 14 days, so a dormant source starts empty
    instead of dumping stale history on top of the feed."""
    return (
        "SELECT GREATEST("
        f"COALESCE((SELECT MIN({id_col}) - 1 FROM "
        f"(SELECT {id_col} FROM {table} ORDER BY {id_col} DESC LIMIT 100) t), 0), "
        f"COALESCE((SELECT MAX({id_col}) FROM {table} "
        f"WHERE {time_col} < NOW() - INTERVAL 14 DAY), 0))"
    )


# Source registry. Each source's SQL must return the normalized columns
# (id, actor, action, entity_type, entity_id, details, created_at); a single
# generic puller does the rest. Add a portal by appending an entry here.
SOURCES = [
    {
        "name": "pkdb.admin_audit_log",
        "portal": "PK",
        "sql": (
            "SELECT id, COALESCE(performed_by, 'Unknown') AS actor, action, "
            "entity_type, entity_id, details, created_at "
            "FROM pkdb.admin_audit_log WHERE id > :wm ORDER BY id LIMIT :lim"
        ),
        "init": _init_sql("pkdb.admin_audit_log"),
    },
    {
        "name": "pkdb.coco_audit_log",
        "portal": "PK",
        "sql": (
            "SELECT id, COALESCE(performed_by, 'Unknown') AS actor, action, "
            "entity_type, entity_id, details, created_at "
            "FROM pkdb.coco_audit_log WHERE id > :wm ORDER BY id LIMIT :lim"
        ),
        "init": _init_sql("pkdb.coco_audit_log"),
    },
    {
        "name": "pkdb.inventory_audit_log",
        "portal": "PK",
        "sql": (
            "SELECT id, COALESCE(performed_by, 'Unknown') AS actor, action, "
            "'item' AS entity_type, item_id AS entity_id, "
            "JSON_OBJECT('item_name', item_name, "
            "'quantity', JSON_OBJECT('from', quantity_before, 'to', quantity_after)"
            ") AS details, created_at "
            "FROM pkdb.inventory_audit_log WHERE id > :wm ORDER BY id LIMIT :lim"
        ),
        "init": _init_sql("pkdb.inventory_audit_log"),
    },
    {
        "name": "financedb.admin_audit_log",
        "portal": "FIN",
        "sql": (
            "SELECT id, COALESCE(performed_by, 'Unknown') AS actor, action, "
            "entity_type, entity_id, details, created_at "
            "FROM financedb.admin_audit_log WHERE id > :wm ORDER BY id LIMIT :lim"
        ),
        "init": _init_sql("financedb.admin_audit_log"),
    },
    {
        "name": "financedb.bill_activity_log",
        "portal": "FIN",
        "sql": (
            "SELECT id, CONCAT('Admin #', COALESCE(performed_by, '?')) AS actor, "
            "LOWER(action) AS action, 'bill' AS entity_type, bill_id AS entity_id, "
            "JSON_OBJECT('status', JSON_OBJECT('from', old_status, 'to', new_status), "
            "'notes', notes) AS details, performed_at AS created_at "
            "FROM financedb.bill_activity_log WHERE id > :wm ORDER BY id LIMIT :lim"
        ),
        "init": _init_sql("financedb.bill_activity_log", time_col="performed_at"),
    },
    {
        # Kouzina Admin Console. Already writes a rich audit_log; we resolve the
        # actor's name off pkdb.admins when it's a kftpl email, parse target's
        # "type:id" into entity, pass meta through as details, and drop the
        # operational noise (auth, tenant switches, AI queries, cookie pastes).
        "name": "kouzinaos.audit_log",
        "portal": "KAC",
        "sql": (
            "SELECT a.id, COALESCE(adm.name, a.user_email, 'Unknown') AS actor, "
            "a.action, "
            "CASE WHEN a.target LIKE '%:%' THEN SUBSTRING_INDEX(a.target,':',1) "
            "ELSE NULL END AS entity_type, "
            "CASE WHEN a.target LIKE '%:%' THEN SUBSTRING_INDEX(a.target,':',-1) "
            "ELSE a.target END AS entity_id, "
            "a.meta AS details, a.created_at "
            "FROM kouzinaos.audit_log a "
            "LEFT JOIN pkdb.admins adm ON adm.email = a.user_email "
            "WHERE a.id > :wm AND a.action NOT IN ("
            "'login','logout','switch_tenant','permission_denied','ai_sql','ai_ask',"
            "'swiggy_cookie_paste','zomato_cookie_paste','petpooja_cookie_paste',"
            "'use_shared_session','super_cron_run_now','password_change') "
            "ORDER BY a.id LIMIT :lim"
        ),
        "init": _init_sql("kouzinaos.audit_log"),
    },
]

# Launch-app actions (kitchen launch: manpower, milestones) live in the PK
# audit log but are their own workstream — badge them LAUNCH, not PK.
LAUNCH_PORTAL = "LAUNCH"
_LAUNCH_MARKERS = ("launch", "milestone", "manpower")


def _portal_for(base_portal: str, entity_type: Optional[str], action: Optional[str]) -> str:
    blob = f"{entity_type or ''} {action or ''}".lower()
    if any(m in blob for m in _LAUNCH_MARKERS):
        return LAUNCH_PORTAL
    return base_portal


# Keep the feed scoped to the registered portals: drop events from sources
# that were removed from the registry (idempotent, runs at boot). "LIVE" is
# this app's own portal (programs, feedback) — never clean it up.
_ACTIVE_PORTALS = sorted({s["portal"] for s in SOURCES} | {"LIVE", LAUNCH_PORTAL})
with SessionLocal() as _db:
    _db.query(FeedEventRow).filter(
        ~FeedEventRow.portal.in_(_ACTIVE_PORTALS)
    ).delete(synchronize_session=False)
    _db.query(IngestStateRow).filter(
        ~IngestStateRow.source.in_([s["name"] for s in SOURCES])
    ).delete(synchronize_session=False)
    # Re-badge already-ingested launch events (they landed as PK before the
    # LAUNCH portal existed). Action names are launch-specific, so this won't
    # touch LIVE program/feedback events.
    _db.query(FeedEventRow).filter(
        FeedEventRow.portal != "LIVE",
        or_(
            FeedEventRow.action.like("%launch%"),
            FeedEventRow.action.like("%milestone%"),
            FeedEventRow.action.like("%manpower%"),
        ),
    ).update({FeedEventRow.portal: LAUNCH_PORTAL}, synchronize_session=False)
    _db.commit()

STATUS_LABELS = {
    "not_started": "Not Started",
    "in_progress": "In Progress",
    "blocked": "Blocked",
    "complete": "Complete",
}


def emit_live_event(actor: str, action: str, summary: str, details: Optional[dict] = None):
    """Publish an in-app event (programs, feedback) to the feed as portal LIVE."""
    with SessionLocal() as db:
        for _ in range(3):  # retry on the rare source_id race between replicas
            try:
                next_id = (
                    db.query(func.max(FeedEventRow.source_id))
                    .filter(FeedEventRow.source == "live")
                    .scalar()
                    or 0
                ) + 1
                db.add(
                    FeedEventRow(
                        portal="LIVE",
                        source="live",
                        source_id=next_id,
                        actor=actor,
                        action=action,
                        summary=summary,
                        details=json.dumps(details) if details else None,
                        happened_at=datetime.now(timezone.utc).replace(tzinfo=None),
                    )
                )
                db.commit()
                return
            except Exception:
                db.rollback()


# Seed the first programs once (idempotent: only when the table is empty).
with SessionLocal() as _db:
    if _db.query(ProgramRow).count() == 0:
        _eta = datetime(2026, 7, 31)
        _db.add(
            ProgramRow(
                name="Use of retort instead of frozen food",
                eta=_eta,
                status="not_started",
                created_by="system",
            )
        )
        _db.add(
            ProgramRow(
                name="V2 UP Menu",
                owner_email="pawan.kumar@kftpl.com",
                owner_name="Pawan",
                eta=_eta,
                status="not_started",
                created_by="system",
            )
        )
        _db.commit()


def _seed_dev_events(db):
    """Local dev (SQLite): fabricate a few events so the UI is testable."""
    if db.query(FeedEventRow).count() > 0:
        return
    now = datetime.now(timezone.utc)
    samples = [
        ("Rohan", "item_setting_changed", "changed settings on item #231"),
        ("ali", "price_update", "price update item #88"),
        ("Priya", "user_create", "created user #14"),
        ("Rohan", "expense_added", "expense added expense #501"),
        # A burst by the same actor — exercises grouping in dev.
        ("Flame & Feast", "stock_deducted", "stock deducted · Manchurian Sauce (98 → 93)"),
        ("Flame & Feast", "stock_deducted", "stock deducted · Coated Paneer (27 → 25)"),
        ("Flame & Feast", "stock_deducted", "stock deducted · Butter Chicken Gravy (206 → 196)"),
        ("Flame & Feast", "stock_deducted", "stock deducted · Egg Corn Fried Rice (743 → 733)"),
        # Different action, same person, same window — exercises mixed grouping.
        ("Flame & Feast", "order_status_changed", "order status changed order #38010"),
    ]
    details = json.dumps(
        {"item_name": "LACCHA PARATHA MAIDA", "active": {"from": True, "to": False}}
    )
    for i, (actor, action, summary) in enumerate(samples):
        db.add(
            FeedEventRow(
                portal="PK",
                source="dev",
                source_id=i + 1,
                actor=actor,
                action=action,
                summary=summary,
                details=details if action == "item_setting_changed" else None,
                happened_at=now - timedelta(minutes=7 * (len(samples) - i)),
            )
        )
    db.commit()


def maybe_ingest():
    """Ingest new audit rows if the last run is stale. Cheap no-op otherwise."""
    if not IS_MYSQL:
        with SessionLocal() as db:
            _seed_dev_events(db)
        return

    now = datetime.now(timezone.utc)
    with SessionLocal() as db:
        states = {s.source: s for s in db.query(IngestStateRow).all()}
        freshest = max(
            (s.last_run for s in states.values() if s.last_run), default=None
        )
        if freshest and now - freshest.replace(tzinfo=timezone.utc) < INGEST_INTERVAL:
            return

    with engine.connect() as conn:
        # Advisory lock: if another replica is ingesting, skip — it'll be fresh.
        got = conn.execute(text("SELECT GET_LOCK('kouzina_feed_ingest', 0)")).scalar()
        if got != 1:
            return
        try:
            with SessionLocal() as db:
                for source in SOURCES:
                    state = db.get(IngestStateRow, source["name"])
                    if state is None:
                        start = conn.execute(text(source["init"])).scalar() or 0
                        state = IngestStateRow(source=source["name"], last_id=start)
                        db.add(state)
                    try:
                        rows = conn.execute(
                            text(source["sql"]),
                            {"wm": state.last_id or 0, "lim": INGEST_BATCH},
                        ).mappings().all()
                    except Exception:
                        # A source we can't read (missing grant, dropped table)
                        # must not take down the whole feed.
                        continue
                    watermark = state.last_id or 0
                    # One query to find already-ingested rows instead of one
                    # SELECT per row — matters now that PK logs busily.
                    ids = [r["id"] for r in rows]
                    existing = set()
                    if ids:
                        existing = {
                            sid
                            for (sid,) in db.query(FeedEventRow.source_id).filter(
                                FeedEventRow.source == source["name"],
                                FeedEventRow.source_id.in_(ids),
                            )
                        }
                    for r in rows:
                        watermark = r["id"]
                        if r["action"] in EXCLUDED_ACTIONS or r["id"] in existing:
                            continue
                        db.add(
                            FeedEventRow(
                                portal=_portal_for(
                                    source["portal"], r["entity_type"], r["action"]
                                ),
                                source=source["name"],
                                source_id=r["id"],
                                actor=r["actor"] or "Unknown",
                                action=r["action"],
                                summary=_summarize(
                                    r["action"], r["entity_type"], r["entity_id"]
                                ),
                                details=r["details"],
                                happened_at=r["created_at"],
                            )
                        )
                    state.last_id = watermark
                    state.last_run = now
                db.commit()
        finally:
            conn.execute(text("SELECT RELEASE_LOCK('kouzina_feed_ingest')"))


# ---- Schemas ---------------------------------------------------------------

class LoginIn(BaseModel):
    email: str = Field(..., min_length=3, max_length=255)
    password: str = Field(..., min_length=1, max_length=255)


class MessageIn(BaseModel):
    recipient: str = Field(..., min_length=3, max_length=255)
    body: str = Field(..., min_length=1, max_length=2000)


class Message(BaseModel):
    id: int
    sender: str
    recipient: str
    body: str
    created_at: str


class CommentIn(BaseModel):
    body: str = Field(..., min_length=1, max_length=1000)


def _serialize_message(r: MessageRow) -> Message:
    return Message(
        id=r.id,
        sender=r.sender,
        recipient=r.recipient,
        body=r.body,
        created_at=_iso_utc(r.created_at),
    )


# ---- App -------------------------------------------------------------------

app = FastAPI(title="Kouzina Live Updates")
api = FastAPI()

for sub in (app, api):
    sub.add_middleware(
        CORSMiddleware,
        allow_origins=["*"],
        allow_methods=["*"],
        allow_headers=["*"],
    )


@app.middleware("http")
async def _no_cache_html(request, call_next):
    """Never let the SPA shell (index.html) be cached — hashed JS/CSS assets
    stay cacheable, but a fresh deploy's index.html must always be re-fetched so
    clients pick up the new asset hashes instead of a stale build."""
    resp = await call_next(request)
    if resp.headers.get("content-type", "").startswith("text/html"):
        resp.headers["Cache-Control"] = "no-cache, no-store, must-revalidate"
    return resp


@api.get("/health")
def health():
    return {"status": "ok"}


@api.post("/login")
def login(body: LoginIn):
    admin = fetch_admin_by_email(body.email.strip())
    if (
        not admin
        or not admin.get("active")
        or not _verify_password(body.password, admin["password_hash"])
    ):
        raise HTTPException(status_code=401, detail="Invalid email or password")
    try:  # usage tracking for the adoption dashboard; never blocks login
        with SessionLocal() as db:
            db.add(LiveLoginRow(email=admin["email"], name=admin.get("name")))
            db.commit()
    except Exception:
        pass
    return {
        "token": make_token(admin["email"], admin.get("name")),
        "email": admin["email"],
        "name": admin.get("name") or admin["email"],
        "is_super": is_superadmin(admin["email"]),
    }


@api.get("/me")
def me(admin: dict = Depends(current_admin)):
    """Fresh identity: name from pkdb (renames show without re-login), org
    profile (department/function) and the superadmin flag."""
    out = {"email": admin["email"], "name": admin.get("name") or admin["email"]}
    account = fetch_admin_by_email(admin["email"])
    if account and account.get("name"):
        out["name"] = account["name"]
    with SessionLocal() as db:
        p = (
            db.query(UserProfileRow)
            .filter(func.lower(UserProfileRow.email) == admin["email"].lower())
            .first()
        )
        out["department"] = p.department if p else None
        out["function"] = p.function if p else None
    out["is_super"] = is_superadmin(admin["email"])
    return out


def _owner_matches(owner: str, name: str) -> bool:
    """Does this display name belong to the owner? Owners in the sheet are short
    names ("Shanil", "GG"), so match exact, first-name or initials."""
    o, n = (owner or "").strip().lower(), (name or "").strip().lower()
    if not o or not n:
        return False
    if n == o or n.startswith(o + " "):
        return True
    parts = n.split()
    if parts and parts[0] == o:
        return True
    return "".join(w[0] for w in parts if w) == o


def _resolve_owner(owner: str) -> Optional[dict]:
    """Find the admin account behind an owner short-name, if unambiguous."""
    if not owner:
        return None
    hits = [a for a in list_active_admins() if _owner_matches(owner, a.get("name") or "")]
    return hits[0] if len(hits) == 1 else None


@api.get("/org/team")
def org_team(
    department: Optional[str] = Query(default=None),
    admin: dict = Depends(current_admin),
):
    """Members of a department (default: mine) — any admin can see their team."""
    with SessionLocal() as db:
        if not department:
            p = (
                db.query(UserProfileRow)
                .filter(func.lower(UserProfileRow.email) == admin["email"].lower())
                .first()
            )
            department = p.department if p else None
        if not department:
            return {"department": None, "members": []}
        profs = (
            db.query(UserProfileRow)
            .filter(UserProfileRow.department == department)
            .all()
        )
        dept_row = (
            db.query(OrgDeptRow)
            .filter(
                OrgDeptRow.active == 1,
                func.lower(OrgDeptRow.department) == department.lower(),
            )
            .first()
        )
    owner = (dept_row.owner if dept_row else None) or next(
        (p.owner for p in profs if p.owner), None
    )
    names = resolve_names({p.email for p in profs})
    members = [
        {
            "name": names.get(p.email) or p.email.split("@")[0],
            "email": p.email,
            "function": p.function,
            "sub_department": p.sub_department,
            "is_owner": bool(owner) and _owner_matches(owner, names.get(p.email) or ""),
        }
        for p in profs
    ]
    # The owner belongs to the team even when their home profile sits in
    # another department (e.g. RK owns Supply Chain + 3rd Party Brands).
    if owner and not any(m["is_owner"] for m in members):
        acct = _resolve_owner(owner)
        members.append(
            {
                "name": (acct or {}).get("name") or owner,
                "email": (acct or {}).get("email"),
                "function": None,
                "sub_department": None,
                "is_owner": True,
            }
        )
    members.sort(key=lambda m: (not m["is_owner"], m["name"].lower()))
    return {"department": department, "owner": owner, "members": members}


@api.get("/admins")
def admins(admin: dict = Depends(current_admin)):
    """All active admins (including you — the DM composer hides self client-side,
    but you should be selectable for red-flag reminders / testing)."""
    return list_active_admins()


# ---- User management (superadmin) ------------------------------------------------
# Accounts live in pkdb.admins (same auth KPK uses); KLU adds the org profile
# (function/department) in its own DB and never touches KPK permissions.

DEFAULT_PASSWORD = "Welcome@123"


def _hash_password(pw: str) -> str:
    """bcrypt with the $2y$ prefix Laravel/PK expects (PHP password_verify
    treats $2y$/$2b$ identically; our own login normalises back)."""
    h = bcrypt.hashpw(pw.encode(), bcrypt.gensalt()).decode()
    return "$2y$" + h[4:] if h.startswith("$2b$") else h


def is_superadmin(email: str) -> bool:
    if (email or "").lower() == SUPERADMIN_EMAIL.lower():
        if IS_MYSQL:  # a deactivated bootstrap account loses its powers too
            try:
                with engine.connect() as conn:
                    row = conn.execute(
                        text("SELECT active FROM pkdb.admins WHERE LOWER(email) = :e LIMIT 1"),
                        {"e": (email or "").lower()},
                    ).first()
                if row is not None and not row[0]:
                    return False
            except Exception:
                pass
        return True
    if IS_MYSQL:
        try:
            with engine.connect() as conn:
                return (
                    conn.execute(
                        text("SELECT 1 FROM pkdb.admins WHERE email = :e "
                             "AND active = 1 AND is_super_admin = 1 LIMIT 1"),
                        {"e": email},
                    ).first()
                    is not None
                )
        except Exception:
            return False
    return False


def require_superadmin(admin: dict = Depends(current_admin)) -> dict:
    if not is_superadmin(admin["email"]):
        raise HTTPException(status_code=403, detail="Superadmin only")
    return admin


def _grant_hint(exc: Exception) -> str:
    msg = str(exc)
    if "denied" in msg.lower() or "1142" in msg:
        return ("The app's DB user lacks write access on pkdb.admins — ask the DBA for: "
                "GRANT SELECT, INSERT, UPDATE ON pkdb.admins TO <app user>;")
    return "Database error while writing pkdb.admins"


def _norm_email(e: str) -> str:
    return (e or "").strip().lower()


def _guard_target(email: str, me: str):
    """Mutations may not touch system accounts or OTHER superadmins — those are
    KPK's to manage. Raises 404/403; returns the admin row (or None on sqlite)."""
    if _excluded(email):
        raise HTTPException(status_code=403, detail="System accounts can't be managed here")
    if not IS_MYSQL:
        return None
    with engine.connect() as conn:
        row = conn.execute(
            text("SELECT email, is_super_admin FROM pkdb.admins WHERE LOWER(email) = :e LIMIT 1"),
            {"e": email},
        ).mappings().first()
    if row is None:
        raise HTTPException(status_code=404, detail="No admin with this email")
    if row["is_super_admin"] and email != _norm_email(me):
        raise HTTPException(
            status_code=403, detail="Superadmin accounts are managed in KPK, not here"
        )
    return dict(row)


class UserIn(BaseModel):
    name: str = Field(..., min_length=2, max_length=255)
    email: str = Field(..., min_length=5, max_length=255)
    function: Optional[str] = Field(default=None, max_length=255)
    department: Optional[str] = Field(default=None, max_length=255)
    sub_department: Optional[str] = Field(default=None, max_length=255)
    owner: Optional[str] = Field(default=None, max_length=255)
    notes: Optional[str] = Field(default=None, max_length=1000)
    password: Optional[str] = Field(default=None, min_length=6, max_length=72)


class UserPatch(BaseModel):
    name: Optional[str] = Field(default=None, min_length=2, max_length=255)
    active: Optional[bool] = None
    function: Optional[str] = Field(default=None, max_length=255)
    department: Optional[str] = Field(default=None, max_length=255)
    sub_department: Optional[str] = Field(default=None, max_length=255)
    owner: Optional[str] = Field(default=None, max_length=255)
    notes: Optional[str] = Field(default=None, max_length=1000)


class PasswordResetIn(BaseModel):
    email: str
    password: Optional[str] = Field(default=None, min_length=6, max_length=72)


class MyPasswordIn(BaseModel):
    current: str = Field(..., min_length=1, max_length=128)
    new: str = Field(..., min_length=6, max_length=72)


def _profiles_by_email(db) -> dict:
    return {p.email.lower(): p for p in db.query(UserProfileRow).all()}


def _apply_profile(db, email: str, data: dict, who: str):
    """Upsert the KLU org profile for an email (only the provided keys)."""
    row = db.query(UserProfileRow).filter(func.lower(UserProfileRow.email) == email.lower()).first()
    if not row:
        row = UserProfileRow(email=email)
        db.add(row)
    for k in ("function", "department", "sub_department", "owner", "notes"):
        if k in data and data[k] is not None:
            setattr(row, k, str(data[k]).strip() or None)
    row.created_by = row.created_by or who
    row.updated_at = datetime.now(timezone.utc).replace(tzinfo=None)
    db.flush()  # visible to later lookups in this same session


@api.get("/org/users")
def org_users(admin: dict = Depends(require_superadmin)):
    """Every pkdb admin + their KLU org profile, for the Users tab."""
    accounts = []
    if IS_MYSQL:
        with engine.connect() as conn:
            rows = conn.execute(
                text("SELECT id, name, email, active, is_super_admin, "
                     "(allowed_modules IS NOT NULL AND allowed_modules <> '' "
                     " AND allowed_modules <> '[]') AS has_kpk "
                     "FROM pkdb.admins ORDER BY active DESC, name ASC")
            ).mappings().all()
            accounts = [dict(r) for r in rows if not _excluded(r["email"])]
    else:  # dev fallback
        accounts = [
            {"id": i + 1, "name": a["name"], "email": a["email"], "active": 1,
             "is_super_admin": 0, "has_kpk": 1}
            for i, a in enumerate(list_active_admins())
        ]
    with SessionLocal() as db:
        profs = _profiles_by_email(db)
    out = []
    for a in accounts:
        p = profs.get(_norm_email(a["email"]))
        out.append(
            {
                "name": a["name"],
                "email": a["email"],
                "active": bool(a["active"]),
                "is_super_admin": bool(a["is_super_admin"]),
                "kpk_access": bool(a["has_kpk"]),
                "function": p.function if p else None,
                "department": p.department if p else None,
                "sub_department": p.sub_department if p else None,
                "owner": p.owner if p else None,
                "notes": p.notes if p else None,
            }
        )
    return {"users": out, "default_password": DEFAULT_PASSWORD}


@api.post("/org/users", status_code=201)
def org_create_user(payload: UserIn, admin: dict = Depends(require_superadmin)):
    email = _norm_email(payload.email)
    if "@" not in email:
        raise HTTPException(status_code=400, detail="Invalid email")
    if _excluded(email):
        raise HTTPException(status_code=403, detail="System accounts can't be managed here")
    name = payload.name.strip()
    if not IS_MYSQL:
        raise HTTPException(status_code=400, detail="User creation needs the production DB")
    with engine.connect() as conn:
        exists = conn.execute(
            text("SELECT 1 FROM pkdb.admins WHERE LOWER(email) = :e LIMIT 1"), {"e": email}
        ).first()
        if exists:
            raise HTTPException(status_code=409, detail="An admin with this email already exists")
        try:
            conn.execute(
                text("INSERT INTO pkdb.admins (name, email, password_hash, is_super_admin, "
                     "active, allowed_modules, created_at) "
                     "VALUES (:n, :e, :p, 0, 1, '[]', NOW())"),
                {"n": name, "e": email, "p": _hash_password(payload.password or DEFAULT_PASSWORD)},
            )
            conn.commit()
        except HTTPException:
            raise
        except Exception as exc:
            raise HTTPException(status_code=500, detail=_grant_hint(exc))
    with SessionLocal() as db:
        _apply_profile(db, email, payload.model_dump(), admin["email"])
        db.commit()
    emit_live_event(
        admin.get("name") or admin["email"], "user_created",
        f"created user · {name}", {"module": "Users", "department": payload.department},
    )
    return {"ok": True, "email": email, "password_set": bool(payload.password)}


@api.patch("/org/users")
def org_update_user(
    email: str, payload: UserPatch, admin: dict = Depends(require_superadmin)
):
    email = _norm_email(email)
    _guard_target(email, admin["email"])
    if IS_MYSQL and (payload.name is not None or payload.active is not None):
        sets, params = [], {"e": email}
        if payload.name is not None:
            sets.append("name = :n")
            params["n"] = payload.name.strip()
        if payload.active is not None:
            if not payload.active and email == _norm_email(admin["email"]):
                raise HTTPException(status_code=400, detail="You can't deactivate yourself")
            sets.append("active = :a")
            params["a"] = 1 if payload.active else 0
        try:
            with engine.connect() as conn:
                res = conn.execute(
                    text(f"UPDATE pkdb.admins SET {', '.join(sets)} WHERE LOWER(email) = :e"),
                    params,
                )
                conn.commit()
                if res.rowcount == 0:
                    raise HTTPException(status_code=404, detail="No admin with this email")
        except HTTPException:
            raise
        except Exception as exc:
            raise HTTPException(status_code=500, detail=_grant_hint(exc))
    with SessionLocal() as db:
        _apply_profile(db, email, payload.model_dump(exclude_none=True), admin["email"])
        db.commit()
    if payload.active is not None:
        emit_live_event(
            admin.get("name") or admin["email"],
            "user_deactivated" if not payload.active else "user_reactivated",
            ("deactivated" if not payload.active else "reactivated") + f" user · {email}",
            {"module": "Users"},
        )
    return {"ok": True}


@api.post("/org/users/password")
def org_reset_password(payload: PasswordResetIn, admin: dict = Depends(require_superadmin)):
    email = _norm_email(payload.email)
    _guard_target(email, admin["email"])
    if not IS_MYSQL:
        raise HTTPException(status_code=400, detail="Password reset needs the production DB")
    try:
        with engine.connect() as conn:
            res = conn.execute(
                text("UPDATE pkdb.admins SET password_hash = :p WHERE LOWER(email) = :e"),
                {"p": _hash_password(payload.password or DEFAULT_PASSWORD), "e": email},
            )
            conn.commit()
            if res.rowcount == 0:
                raise HTTPException(status_code=404, detail="No admin with this email")
    except HTTPException:
        raise
    except Exception as exc:
        raise HTTPException(status_code=500, detail=_grant_hint(exc))
    emit_live_event(
        admin.get("name") or admin["email"], "user_password_reset",
        f"reset password · {email}", {"module": "Users"},
    )
    return {"ok": True, "default": payload.password is None}


@api.post("/me/password")
def change_my_password(payload: MyPasswordIn, admin: dict = Depends(current_admin)):
    """Any admin can change their own password after proving the current one."""
    account = fetch_admin_by_email(admin["email"])
    if not account or not _verify_password(payload.current, account["password_hash"]):
        raise HTTPException(status_code=400, detail="Current password is incorrect")
    if not IS_MYSQL:
        raise HTTPException(status_code=400, detail="Password change needs the production DB")
    try:
        with engine.connect() as conn:
            conn.execute(
                text("UPDATE pkdb.admins SET password_hash = :p WHERE LOWER(email) = :e"),
                {"p": _hash_password(payload.new), "e": _norm_email(admin["email"])},
            )
            conn.commit()
    except Exception as exc:
        raise HTTPException(status_code=500, detail=_grant_hint(exc))
    return {"ok": True}


# ---- Org structure (functions & departments) -------------------------------------

class OrgDeptIn(BaseModel):
    function: str = Field(..., min_length=2, max_length=255)
    department: str = Field(..., min_length=2, max_length=255)
    leader: Optional[str] = Field(default=None, max_length=255)
    owner: Optional[str] = Field(default=None, max_length=255)


@api.get("/org/structure")
def org_structure(admin: dict = Depends(current_admin)):
    """Org chart = UNION of curated structure rows and departments that exist
    only on people's profiles (so a people-sheet department like "Partner
    Relationship & Program Management" still shows, with function/owner
    inferred from its members). Every row carries a member count. Readable by
    all admins; editing is superadmin-only."""
    from collections import Counter, defaultdict

    with SessionLocal() as db:
        rows = db.query(OrgDeptRow).filter(OrgDeptRow.active == 1).all()
        profs = (
            db.query(
                UserProfileRow.email, UserProfileRow.department,
                UserProfileRow.function, UserProfileRow.owner,
            )
            .filter(UserProfileRow.department.isnot(None))
            .all()
        )
        prog_counts_raw = dict(
            db.query(ProgramRow.department, func.count())
            .filter(ProgramRow.active == 1, ProgramRow.department.isnot(None))
            .group_by(ProgramRow.department)
            .all()
        )
    prog_counts = {(k or "").strip().lower(): v for k, v in prog_counts_raw.items()}
    member_names = resolve_names({p.email for p in profs})
    counts: Counter = Counter()
    fn_by, own_by, canonical = defaultdict(Counter), defaultdict(Counter), {}
    names_by = defaultdict(list)
    for email, dept, fn, own in profs:
        key = dept.strip().lower()
        counts[key] += 1
        names_by[key].append(member_names.get(email) or email.split("@")[0])
        canonical.setdefault(key, dept.strip())
        if fn:
            fn_by[key][fn.strip()] += 1
        if own:
            own_by[key][own.strip()] += 1

    def _count(key: str, owner: Optional[str]) -> int:
        # Owners belong to their department's team even when their home
        # profile points elsewhere (RK → Supply Chain AND 3rd Party Brands).
        base = counts.get(key, 0)
        if owner and not any(_owner_matches(owner, n) for n in names_by.get(key, [])):
            base += 1
        return base

    out, seen = [], set()
    for r in rows:
        key = r.department.strip().lower()
        seen.add(key)
        out.append(
            {"id": r.id, "function": r.function, "department": r.department,
             "leader": r.leader, "owner": r.owner, "members": _count(key, r.owner),
             "programs": prog_counts.get(key, 0)}
        )
    for key, cnt in counts.items():
        if key in seen:
            continue
        own = own_by[key].most_common(1)[0][0] if own_by[key] else None
        out.append(
            {"id": None,
             "function": fn_by[key].most_common(1)[0][0] if fn_by[key] else None,
             "department": canonical[key], "leader": None,
             "owner": own, "members": _count(key, own),
             "programs": prog_counts.get(key, 0)}
        )
    out.sort(key=lambda r: r["department"].lower())
    return {
        "rows": out,
        "functions": sorted({r["function"] for r in out if r["function"]}),
        "departments": sorted({r["department"] for r in out}),
        "is_super": is_superadmin(admin["email"]),
    }


@api.post("/org/structure", status_code=201)
def org_structure_add(payload: OrgDeptIn, admin: dict = Depends(require_superadmin)):
    with SessionLocal() as db:
        row = OrgDeptRow(
            function=payload.function.strip(), department=payload.department.strip(),
            leader=(payload.leader or "").strip() or None,
            owner=(payload.owner or "").strip() or None,
        )
        db.add(row)
        db.commit()
        return {"ok": True, "id": row.id}


@api.patch("/org/structure/{row_id}")
def org_structure_edit(row_id: int, payload: OrgDeptIn, admin: dict = Depends(require_superadmin)):
    with SessionLocal() as db:
        row = db.get(OrgDeptRow, row_id)
        if not row:
            raise HTTPException(status_code=404, detail="Not found")
        row.function = payload.function.strip()
        row.department = payload.department.strip()
        row.leader = (payload.leader or "").strip() or None
        row.owner = (payload.owner or "").strip() or None
        db.commit()
    return {"ok": True}


@api.delete("/org/structure/{row_id}")
def org_structure_remove(row_id: int, admin: dict = Depends(require_superadmin)):
    with SessionLocal() as db:
        row = db.get(OrgDeptRow, row_id)
        if row:
            row.active = 0
            db.commit()
    return {"ok": True}


# ---- Excel import (ownership.xlsx) ------------------------------------------------

def _sheet_header_map(header_row) -> dict:
    """Tolerant header → key mapping ('Sub Department' → sub_department, …)."""
    keys = {}
    for idx, cell in enumerate(header_row):
        label = str(cell or "").strip().lower()
        if not label:
            continue
        if "person" in label or label == "name":
            keys["name"] = idx
        elif "email" in label:
            keys["email"] = idx
        elif "sub" in label and "depart" in label:
            keys["sub_department"] = idx
        elif "function" in label:
            keys["function"] = idx
        elif "owner" in label:
            keys["owner"] = idx
        elif "leader" in label:
            keys["leader"] = idx
        elif "depart" in label:
            keys["department"] = idx
        elif "note" in label:
            keys["notes"] = idx
    return keys


@api.post("/org/import")
async def org_import(
    file: UploadFile = File(...),
    apply: bool = Query(default=False),
    admin: dict = Depends(require_superadmin),
):
    """Import ownership.xlsx: people sheet (Person/Email/Function/Department/…)
    and an optional departments sheet (Function/Department/Leader/Owner).
    Dry-run by default — returns the plan; ?apply=true executes it: creates
    missing pkdb admins (default password, no KPK modules), upserts profiles
    and the org structure."""
    try:
        from openpyxl import load_workbook
        wb = load_workbook(io.BytesIO(await file.read()), data_only=True)
    except Exception:
        raise HTTPException(status_code=400, detail="Could not read that file — is it .xlsx?")

    people, depts, skipped = [], [], []
    for ws in wb.worksheets:
        rows = list(ws.iter_rows(values_only=True))
        header_idx = next(
            (i for i, r in enumerate(rows)
             if r and _sheet_header_map(r).keys() >= {"department"}),
            None,
        )
        if header_idx is None:
            continue
        keys = _sheet_header_map(rows[header_idx])
        get = lambda r, k: (str(r[keys[k]]).strip() if k in keys and keys[k] < len(r) and r[keys[k]] is not None else None)
        for r in rows[header_idx + 1:]:
            if not r or not any(v is not None for v in r):
                continue
            if "email" in keys and get(r, "email"):
                email = _norm_email(get(r, "email"))
                if "@" not in email or "." not in email.split("@")[-1] or _excluded(email):
                    skipped.append(get(r, "email"))
                    continue
                people.append({
                    "name": get(r, "name"), "email": email,
                    "function": get(r, "function"), "department": get(r, "department"),
                    "sub_department": get(r, "sub_department"), "owner": get(r, "owner"),
                    "notes": get(r, "notes"),
                })
            elif "email" not in keys and get(r, "department") and get(r, "function"):
                depts.append({
                    "function": get(r, "function"), "department": get(r, "department"),
                    "leader": get(r, "leader"), "owner": get(r, "owner"),
                })

    # Last row wins on duplicates — keeps the upserts single-shot per key.
    people = list({p["email"]: p for p in people}.values())
    depts = list({(d["function"].lower(), d["department"].lower()): d for d in depts}.values())

    if not people and not depts:
        raise HTTPException(status_code=400, detail="No recognisable rows found in the file")

    existing = {}
    if IS_MYSQL:
        with engine.connect() as conn:
            existing = {
                r["email"].lower(): dict(r)
                for r in conn.execute(
                    text("SELECT email, name, active FROM pkdb.admins")
                ).mappings().all()
            }
    else:
        existing = {a["email"].lower(): a for a in list_active_admins()}

    to_create = [p for p in people if p["email"] not in existing]
    to_match = [p for p in people if p["email"] in existing]

    plan = {
        "people_in_file": len(people),
        "matched_existing": len(to_match),
        "will_create": [
            {"name": p["name"], "email": p["email"], "department": p["department"]}
            for p in to_create
        ],
        "departments_in_file": len(depts),
        "skipped_invalid": skipped,
        "default_password": DEFAULT_PASSWORD,
        "applied": False,
    }
    if not apply:
        return plan

    created, failed = 0, []
    if to_create:
        if not IS_MYSQL:
            raise HTTPException(status_code=400, detail="User creation needs the production DB")
        with engine.connect() as conn:
            for p in to_create:
                try:
                    conn.execute(
                        text("INSERT INTO pkdb.admins (name, email, password_hash, "
                             "is_super_admin, active, allowed_modules, created_at) "
                             "VALUES (:n, :e, :p, 0, 1, '[]', NOW())"),
                        {"n": p["name"] or p["email"].split("@")[0], "e": p["email"],
                         "p": _hash_password(DEFAULT_PASSWORD)},
                    )
                    created += 1
                except Exception as exc:
                    failed.append({"email": p["email"], "error": _grant_hint(exc)})
            conn.commit()

    with SessionLocal() as db:
        for p in people:
            if p["email"] not in {f["email"] for f in failed}:
                _apply_profile(db, p["email"], p, admin["email"])
        for d in depts:
            row = (
                db.query(OrgDeptRow)
                .filter(
                    func.lower(OrgDeptRow.function) == d["function"].lower(),
                    func.lower(OrgDeptRow.department) == d["department"].lower(),
                )
                .first()
            )
            if not row:
                row = OrgDeptRow(function=d["function"], department=d["department"])
                db.add(row)
                db.flush()
            row.leader = d["leader"] or row.leader
            row.owner = d["owner"] or row.owner
            row.active = 1
        db.commit()

    emit_live_event(
        admin.get("name") or admin["email"], "users_imported",
        f"imported org sheet · {created} new users, {len(to_match)} matched",
        {"module": "Users"},
    )
    plan.update({"applied": True, "created": created, "profiles_updated": len(people) - len(failed),
                 "failed": failed})
    return plan


@api.get("/leaderboard")
def leaderboard(admin: dict = Depends(current_admin)):
    """Most active people in the last 12 hours, by number of feed events.
    Each entry resolves to an admin email when the actor name matches an
    active admin — that's what makes them messageable."""
    since = datetime.now(timezone.utc).replace(tzinfo=None) - timedelta(hours=12)
    with SessionLocal() as db:
        rows = (
            db.query(FeedEventRow.actor, func.count())
            .filter(FeedEventRow.happened_at >= since)
            .group_by(FeedEventRow.actor)
            .order_by(func.count().desc())
            .limit(50)
            .all()
        )
    entries = [{"actor": a, "count": c, "email": None} for a, c in rows]
    names = [e["actor"] for e in entries]
    if names:
        if IS_MYSQL:
            with engine.connect() as conn:
                pairs = conn.execute(
                    text(
                        "SELECT name, email FROM pkdb.admins "
                        "WHERE active = 1 AND name IN :names"
                    ).bindparams(bindparam("names", expanding=True)),
                    {"names": names},
                ).all()
            by_name = {n: em for n, em in pairs}
        else:
            by_name = {a["name"]: a["email"] for a in list_active_admins()}
        for e in entries:
            em = by_name.get(e["actor"])
            e["email"] = None if _excluded(em) else em
    return entries


@api.get("/overview")
def overview(admin: dict = Depends(current_admin)):
    """Live-page dashboard widgets: program load per owner, private messages
    awaiting MY reply (grouped by person), and open feedback count."""
    me = admin["email"]
    with SessionLocal() as db:
        owner_rows = (
            db.query(ProgramRow.owner_email, ProgramRow.owner_name, func.count())
            .filter(ProgramRow.active == 1, ProgramRow.owner_email.isnot(None))
            .group_by(ProgramRow.owner_email, ProgramRow.owner_name)
            .all()
        )
        program_owners = sorted(
            ({"email": e, "name": n or e, "count": int(c)} for e, n, c in owner_rows),
            key=lambda o: (-o["count"], o["name"].lower()),
        )

        # Every message is directed at one admin (replies inherit the root's
        # recipient), so a conversation is a directed pair. No read receipts
        # exist, so "responded" = sent a later message back on that thread.
        rows = (
            db.query(MessageRow.id, MessageRow.sender, MessageRow.recipient)
            .order_by(MessageRow.id.asc())
            .all()
        )
        latest_pair: dict[tuple, int] = {}   # (sender, recipient) -> latest id
        my_incoming: dict[str, list] = {}    # sender -> ids they sent me
        for mid, s, r in rows:
            if not s or not r or s == r:
                continue
            latest_pair[(s, r)] = mid  # asc order → last write is the max id
            if r == me:
                my_incoming.setdefault(s, []).append(mid)

        # Org-wide: for each person R, how many distinct people are still
        # waiting on R (R got the last word and never replied).
        owes: dict[str, set] = {}
        for (s, r), mid in latest_pair.items():
            if mid > latest_pair.get((r, s), 0):
                owes.setdefault(r, set()).add(s)
        awaiting = [{"email": r, "count": len(waiters)} for r, waiters in owes.items()]

        # Personal: people I owe a reply to, counted by unanswered messages.
        waiting = []
        for other, ids in my_incoming.items():
            lr = latest_pair.get((me, other), 0)
            cnt = sum(1 for i in ids if i > lr)
            if cnt:
                waiting.append({"email": other, "count": cnt})

        names = resolve_names({x["email"] for x in awaiting} | {w["email"] for w in waiting})
        for x in awaiting:
            x["name"] = names.get(x["email"], x["email"])
        for w in waiting:
            w["name"] = names.get(w["email"], w["email"])
        awaiting.sort(key=lambda x: (-x["count"], x["name"].lower()))
        waiting.sort(key=lambda w: (-w["count"], w["name"].lower()))

        fb_open = db.query(FeedbackRow).filter(FeedbackRow.status == "open").count()

    return {
        "program_owners": program_owners,
        "awaiting_response": awaiting,
        "awaiting_response_total": len(awaiting),
        "messages_waiting": waiting,
        "messages_waiting_total": sum(w["count"] for w in waiting),
        "feedback_open": fb_open,
    }


# ---- Message wall (visible to all admins, Twitter-mentions style) -----------

@api.get("/wall/{email}")
def get_wall(email: str, admin: dict = Depends(current_admin)):
    """Recent PUBLIC messages sent TO this person — shown to all admins."""
    with SessionLocal() as db:
        rows = (
            db.query(MessageRow)
            .filter(MessageRow.recipient == email, MessageRow.is_private == 0)
            .order_by(MessageRow.id.desc())
            .limit(30)
            .all()
        )
    names = resolve_names({r.sender for r in rows})
    with SessionLocal() as db:
        counts, mine = _msg_like_info(db, [r.id for r in rows], admin["email"])
    return [
        {
            "id": r.id,
            "sender": r.sender,
            "sender_name": names.get(r.sender) or r.sender.split("@")[0],
            "body": r.body,
            "created_at": _iso_utc(r.created_at),
            "like_count": counts.get(r.id, 0),
            "liked_by_me": r.id in mine,
        }
        for r in rows
    ]


@api.get("/person/{actor}")
def person(actor: str, admin: dict = Depends(current_admin)):
    """Resolve a feed actor: 12h activity count + admin email if messageable."""
    since = datetime.now(timezone.utc).replace(tzinfo=None) - timedelta(hours=12)
    with SessionLocal() as db:
        count = (
            db.query(FeedEventRow)
            .filter(FeedEventRow.actor == actor, FeedEventRow.happened_at >= since)
            .count()
        )
    email = None
    if IS_MYSQL:
        with engine.connect() as conn:
            r = conn.execute(
                text("SELECT email FROM pkdb.admins WHERE name = :n AND active = 1 LIMIT 1"),
                {"n": actor},
            ).first()
        email = r[0] if r else None
    else:
        email = next(
            (a["email"] for a in list_active_admins() if a["name"] == actor), None
        )
    if _excluded(email):
        email = None
    owner_conds = [ProgramRow.owner_name == actor]
    if email:
        owner_conds.append(ProgramRow.owner_email == email)
    with SessionLocal() as db:
        prog_rows = (
            db.query(ProgramRow)
            .filter(ProgramRow.active == 1, or_(*owner_conds))
            .order_by(ProgramRow.eta.asc())
            .all()
        )
    return {
        "actor": actor,
        "count": count,
        "email": email,
        "programs": [
            {
                "id": p.id,
                "name": p.name,
                "status": p.status,
                "eta": p.eta.date().isoformat() if p.eta else None,
            }
            for p in prog_rows
        ],
    }


# ---- Messages hub (public tree + private) ------------------------------------

def _msg_like_info(db, ids: list[int], me_email: str):
    counts, mine = {}, set()
    if ids:
        for mid, c in (
            db.query(MessageLikeRow.message_id, func.count())
            .filter(MessageLikeRow.message_id.in_(ids))
            .group_by(MessageLikeRow.message_id)
        ):
            counts[mid] = c
        mine = {
            r[0]
            for r in db.query(MessageLikeRow.message_id).filter(
                MessageLikeRow.message_id.in_(ids),
                MessageLikeRow.admin_email == me_email,
            )
        }
    return counts, mine


def _serialize_msgs(db, rows, me_email: str) -> list[dict]:
    emails = {r.sender for r in rows} | {r.recipient for r in rows}
    names = resolve_names(emails)
    counts, mine = _msg_like_info(db, [r.id for r in rows], me_email)
    return [
        {
            "id": r.id,
            "parent_id": r.parent_id,
            "sender": r.sender,
            "sender_name": names.get(r.sender) or r.sender.split("@")[0],
            "recipient": r.recipient,
            "recipient_name": names.get(r.recipient) or r.recipient.split("@")[0],
            "is_private": bool(r.is_private),
            "body": r.body,
            "created_at": _iso_utc(r.created_at),
            "like_count": counts.get(r.id, 0),
            "liked_by_me": r.id in mine,
        }
        for r in rows
    ]


@api.get("/messages/public")
def messages_public(admin: dict = Depends(current_admin)):
    """Everyone's public messages (roots + replies), visible to all admins."""
    with SessionLocal() as db:
        rows = (
            db.query(MessageRow)
            .filter(MessageRow.is_private == 0)
            .order_by(MessageRow.id.desc())
            .limit(150)
            .all()
        )
        return _serialize_msgs(db, rows, admin["email"])


@api.get("/messages/private")
def messages_private(admin: dict = Depends(current_admin)):
    """My private messages only — sender or recipient is me."""
    me_email = admin["email"]
    with SessionLocal() as db:
        rows = (
            db.query(MessageRow)
            .filter(
                MessageRow.is_private == 1,
                or_(MessageRow.sender == me_email, MessageRow.recipient == me_email),
            )
            .order_by(MessageRow.id.desc())
            .limit(150)
            .all()
        )
        return _serialize_msgs(db, rows, me_email)


@api.get("/messages/conversation")
def messages_conversation(email: str, admin: dict = Depends(current_admin)):
    """Everything about one person I'm allowed to see: my full thread with them
    (public + private) PLUS the public chains they're part of. Each message is
    tagged so the UI can show two things — what I owe THEM (`owe_me`) and what
    THEY owe others (`owe_them`, the "not responded" chains). Private threads
    between them and third parties stay hidden (I'm not a party)."""
    me_email = admin["email"]
    other = (email or "").strip()
    with SessionLocal() as db:
        rows = (
            db.query(MessageRow)
            .filter(
                or_(
                    # my own thread with them — public + private
                    and_(MessageRow.sender == me_email, MessageRow.recipient == other),
                    and_(MessageRow.sender == other, MessageRow.recipient == me_email),
                    # public messages they sent or received (visible to everyone)
                    and_(
                        MessageRow.is_private == 0,
                        or_(MessageRow.sender == other, MessageRow.recipient == other),
                    ),
                )
            )
            .order_by(MessageRow.id.desc())
            .limit(400)
            .all()
        )
        msgs = _serialize_msgs(db, rows, me_email)

    # Latest reply from each party, to work out who still owes whom.
    my_last_to_other = 0
    other_last_to: dict[str, int] = {}
    for m in sorted(msgs, key=lambda x: x["id"]):
        if m["sender"] == me_email and m["recipient"] == other:
            my_last_to_other = m["id"]
        if m["sender"] == other:
            other_last_to[m["recipient"]] = m["id"]

    owe_me = owe_them = 0
    for m in msgs:
        # they messaged me and I haven't replied since → I owe them
        m["owe_me"] = m["sender"] == other and m["recipient"] == me_email and m["id"] > my_last_to_other
        # someone messaged them and they haven't replied since → they owe others
        m["owe_them"] = m["recipient"] == other and m["id"] > other_last_to.get(m["sender"], 0)
        owe_me += 1 if m["owe_me"] else 0
        owe_them += 1 if m["owe_them"] else 0
    return {"messages": msgs, "owe_me": owe_me, "owe_them": owe_them}


@api.post("/messages/{message_id}/like")
def toggle_message_like(message_id: int, admin: dict = Depends(current_admin)):
    with SessionLocal() as db:
        if not db.get(MessageRow, message_id):
            raise HTTPException(status_code=404, detail="Message not found")
        existing = (
            db.query(MessageLikeRow)
            .filter_by(message_id=message_id, admin_email=admin["email"])
            .first()
        )
        if existing:
            db.delete(existing)
            liked = False
        else:
            db.add(MessageLikeRow(message_id=message_id, admin_email=admin["email"]))
            liked = True
        db.commit()
        count = db.query(MessageLikeRow).filter_by(message_id=message_id).count()
    return {"liked": liked, "like_count": count}


class SendIn(BaseModel):
    recipient: Optional[str] = Field(default=None, max_length=255)
    body: str = Field(..., min_length=1, max_length=2000)
    private: bool = False
    parent_id: Optional[int] = None


@api.post("/messages/send", status_code=201)
def messages_send(payload: SendIn, admin: dict = Depends(current_admin)):
    text_body = payload.body.strip()
    if not text_body:
        raise HTTPException(status_code=400, detail="Message cannot be empty")
    with SessionLocal() as db:
        if payload.parent_id:
            root = db.get(MessageRow, payload.parent_id)
            if not root:
                raise HTTPException(status_code=404, detail="Thread not found")
            if root.is_private and admin["email"] not in (root.sender, root.recipient):
                raise HTTPException(status_code=403, detail="Not your thread")
            row = MessageRow(
                sender=admin["email"],
                recipient=root.recipient,
                body=text_body,
                is_private=root.is_private,
                parent_id=root.parent_id or root.id,  # keep threads one level deep
            )
        else:
            recipient = (payload.recipient or "").strip()
            if not recipient:
                raise HTTPException(status_code=400, detail="Recipient required")
            if recipient == admin["email"]:
                raise HTTPException(status_code=400, detail="You can't message yourself")
            if not is_active_admin(recipient):
                raise HTTPException(
                    status_code=400, detail="Recipient must be an active admin"
                )
            row = MessageRow(
                sender=admin["email"],
                recipient=recipient,
                body=text_body,
                is_private=1 if payload.private else 0,
            )
        db.add(row)
        db.commit()
    return {"ok": True}


# ---- Superadmin dashboard -----------------------------------------------------

@api.get("/dashboard")
def dashboard(admin: dict = Depends(current_admin)):
    if admin["email"].lower() != SUPERADMIN_EMAIL.lower():
        raise HTTPException(status_code=403, detail="Superadmin only")
    now = datetime.now(timezone.utc).replace(tzinfo=None)
    d7, d14 = now - timedelta(days=7), now - timedelta(days=14)
    with SessionLocal() as db:
        totals = {
            "events_7d": db.query(FeedEventRow)
            .filter(FeedEventRow.happened_at >= d7)
            .count(),
            "people_7d": db.query(FeedEventRow.actor)
            .filter(FeedEventRow.happened_at >= d7)
            .distinct()
            .count(),
            "messages_7d": db.query(MessageRow)
            .filter(MessageRow.created_at >= d7)
            .count(),
            "reactions_7d": db.query(EventLikeRow)
            .filter(EventLikeRow.created_at >= d7)
            .count()
            + db.query(EventCommentRow)
            .filter(EventCommentRow.created_at >= d7)
            .count(),
        }
        by_person = [
            {"actor": a, "count": c}
            for a, c in db.query(FeedEventRow.actor, func.count())
            .filter(FeedEventRow.happened_at >= d7)
            .group_by(FeedEventRow.actor)
            .order_by(func.count().desc())
            .limit(12)
            .all()
        ]
        feed_by_day = [
            {"date": str(d), "count": c}
            for d, c in db.query(func.date(FeedEventRow.happened_at), func.count())
            .filter(FeedEventRow.happened_at >= d14)
            .group_by(func.date(FeedEventRow.happened_at))
            .order_by(func.date(FeedEventRow.happened_at))
            .all()
        ]

    def _logins_per_day(table: str) -> list[dict]:
        if not IS_MYSQL:
            return []
        try:
            with engine.connect() as conn:
                rows = conn.execute(
                    text(
                        f"SELECT DATE(created_at) AS d, COUNT(*) AS c FROM {table} "
                        "WHERE action = 'login' AND created_at >= :since "
                        "GROUP BY DATE(created_at) ORDER BY d"
                    ),
                    {"since": d14},
                ).all()
            return [{"date": str(d), "count": c} for d, c in rows]
        except Exception:
            return []

    # ---- Kouzina Live adoption: who signs in, how often, what they do ------
    d30 = now - timedelta(days=30)
    with SessionLocal() as db:
        adoption = {
            "logins_7d": db.query(LiveLoginRow)
            .filter(LiveLoginRow.created_at >= d7)
            .count(),
            "users_7d": db.query(LiveLoginRow.email)
            .filter(LiveLoginRow.created_at >= d7)
            .distinct()
            .count(),
        }
        users: dict[str, dict] = {}

        def _bucket(email):
            return users.setdefault(
                email,
                {
                    "email": email,
                    "logins": 0,
                    "last_login": None,
                    "messages": 0,
                    "likes": 0,
                    "comments": 0,
                },
            )

        for email, cnt, last in (
            db.query(LiveLoginRow.email, func.count(), func.max(LiveLoginRow.created_at))
            .filter(LiveLoginRow.created_at >= d30)
            .group_by(LiveLoginRow.email)
        ):
            u = _bucket(email)
            u["logins"] = cnt
            u["last_login"] = _iso_utc(last)
        for email, cnt in (
            db.query(MessageRow.sender, func.count())
            .filter(MessageRow.created_at >= d30)
            .group_by(MessageRow.sender)
        ):
            _bucket(email)["messages"] = cnt
        for email, cnt in (
            db.query(EventLikeRow.admin_email, func.count())
            .filter(EventLikeRow.created_at >= d30)
            .group_by(EventLikeRow.admin_email)
        ):
            _bucket(email)["likes"] = cnt
        for email, cnt in (
            db.query(EventCommentRow.admin_email, func.count())
            .filter(EventCommentRow.created_at >= d30)
            .group_by(EventCommentRow.admin_email)
        ):
            _bucket(email)["comments"] = cnt

    names = resolve_names(set(users.keys()))
    for email, u in users.items():
        u["name"] = names.get(email) or email.split("@")[0]
    adoption["users"] = sorted(
        users.values(),
        key=lambda u: (u["logins"], u["messages"] + u["likes"] + u["comments"]),
        reverse=True,
    )
    active_emails = {e for e, u in users.items() if u["logins"] > 0}
    adoption["never"] = sorted(
        a["name"] or a["email"]
        for a in list_active_admins()
        if a["email"] not in active_emails
    )

    return {
        "totals": totals,
        "by_person": by_person,
        "feed_by_day": feed_by_day,
        "pk_usage": _logins_per_day("pkdb.admin_audit_log"),
        "kfc_usage": _logins_per_day("financedb.kfc_access_log"),
        "adoption": adoption,
    }


@api.post("/wall/{email}", status_code=201)
def post_wall(email: str, body: CommentIn, admin: dict = Depends(current_admin)):
    text_body = body.body.strip()
    if not text_body:
        raise HTTPException(status_code=400, detail="Message cannot be empty")
    if email == admin["email"]:
        raise HTTPException(status_code=400, detail="You can't message yourself")
    if not is_active_admin(email):
        raise HTTPException(status_code=400, detail="Recipient must be an active admin")
    with SessionLocal() as db:
        row = MessageRow(sender=admin["email"], recipient=email, body=text_body)
        db.add(row)
        db.commit()
    return {"ok": True}


# ---- Live Updates feed ------------------------------------------------------

@api.get("/feed")
def get_feed(
    background_tasks: BackgroundTasks,
    limit: int = Query(default=10, le=100),
    cursor_ts: Optional[str] = Query(default=None),
    cursor_id: Optional[int] = Query(default=None),
    portal: Optional[str] = Query(default=None),
    actor: Optional[str] = Query(default=None),
    department: Optional[str] = Query(default=None),
    admin: dict = Depends(current_admin),
):
    """Latest activity first, GROUPED: consecutive events by the same person
    doing the same action within GROUP_WINDOW collapse into one card (so a
    burst of per-item stock deductions reads as a single update). The first
    page responds immediately from the local table; ingestion of new audit
    rows runs AFTER the response is sent. `limit` counts groups; page into the
    past with the returned next_cursor_ts / next_cursor_id."""
    if cursor_ts is None:
        background_tasks.add_task(maybe_ingest)
    with SessionLocal() as db:
        q = db.query(FeedEventRow)
        if portal:
            q = q.filter(FeedEventRow.portal == portal)
        if actor:
            q = q.filter(FeedEventRow.actor == actor)
        if department:
            # Department → member emails (profiles) → display names (pkdb) →
            # actor filter. Actors in the feed are display names.
            emails = [
                e for (e,) in db.query(UserProfileRow.email)
                .filter(UserProfileRow.department == department)
            ]
            names = set(resolve_names(set(emails)).values())
            # People whose pkdb name is missing still match by email prefix.
            names |= {e.split("@")[0] for e in emails}
            if not names:
                return {"events": [], "has_more": False,
                        "next_cursor_ts": None, "next_cursor_id": None}
            q = q.filter(FeedEventRow.actor.in_(sorted(names)))
        if cursor_ts is not None and cursor_id is not None:
            try:
                ts = datetime.fromisoformat(cursor_ts)
            except ValueError:
                raise HTTPException(status_code=400, detail="Bad cursor")
            if ts.tzinfo is not None:
                # Columns hold naive UTC; compare like with like.
                ts = ts.astimezone(timezone.utc).replace(tzinfo=None)
            q = q.filter(
                or_(
                    FeedEventRow.happened_at < ts,
                    and_(
                        FeedEventRow.happened_at == ts, FeedEventRow.id < cursor_id
                    ),
                )
            )
        raw_fetch = min(limit * 12, 300)
        rows = (
            q.order_by(FeedEventRow.happened_at.desc(), FeedEventRow.id.desc())
            .limit(raw_fetch + 1)
            .all()
        )
        overflow = len(rows) > raw_fetch
        rows = rows[:raw_fetch]

        # Collapse consecutive events by the same person (any actions) within
        # GROUP_WINDOW into one card — a burst of PO generate/push pairs reads
        # as a single update.
        groups: list[dict] = []
        consumed = 0
        for r in rows:
            g = groups[-1] if groups else None
            r_ts = r.happened_at or datetime.min
            if (
                g is not None
                and g["portal"] == r.portal
                and g["actor"] == r.actor
                and (g["anchor"] - r_ts) <= GROUP_WINDOW
            ):
                g["members"].append(r)
                g["actions"].add(r.action)
                consumed += 1
                continue
            if len(groups) == limit:
                break
            groups.append(
                {
                    "portal": r.portal,
                    "actor": r.actor,
                    "anchor": r_ts,
                    "members": [r],
                    "actions": {r.action},
                }
            )
            consumed += 1

        has_more = overflow or consumed < len(rows)
        last = rows[consumed - 1] if consumed else None

        rep_ids = [g["members"][0].id for g in groups]
        like_counts: dict[int, int] = {}
        my_likes: set[int] = set()
        comment_counts: dict[int, int] = {}
        if rep_ids:
            for event_id, cnt in (
                db.query(EventLikeRow.event_id, func.count())
                .filter(EventLikeRow.event_id.in_(rep_ids))
                .group_by(EventLikeRow.event_id)
            ):
                like_counts[event_id] = cnt
            my_likes = {
                r[0]
                for r in db.query(EventLikeRow.event_id).filter(
                    EventLikeRow.event_id.in_(rep_ids),
                    EventLikeRow.admin_email == admin["email"],
                )
            }
            for event_id, cnt in (
                db.query(EventCommentRow.event_id, func.count())
                .filter(EventCommentRow.event_id.in_(rep_ids))
                .group_by(EventCommentRow.event_id)
            ):
                comment_counts[event_id] = cnt

    events = []
    for g in groups:
        rep = g["members"][0]  # newest member represents the group
        events.append(
            {
                "id": rep.id,
                "portal": rep.portal,
                "actor": rep.actor,
                "action": rep.action,
                "actions": sorted(g["actions"])[:6],
                "uniform": len(g["actions"]) == 1,
                "summary": _render_summary(rep.action, rep.summary, rep.details),
                "happened_at": _iso_utc(rep.happened_at),
                "count": len(g["members"]),
                "extras": [
                    _render_summary(m.action, m.summary, m.details)
                    for m in g["members"][1:6]
                ],
                "like_count": like_counts.get(rep.id, 0),
                "liked_by_me": rep.id in my_likes,
                "comment_count": comment_counts.get(rep.id, 0),
            }
        )
    return {
        "events": events,
        "has_more": has_more,
        "next_cursor_ts": _iso_utc(last.happened_at) if last else None,
        "next_cursor_id": last.id if last else None,
    }


@api.post("/feed/{event_id}/like")
def toggle_like(event_id: int, admin: dict = Depends(current_admin)):
    with SessionLocal() as db:
        if not db.get(FeedEventRow, event_id):
            raise HTTPException(status_code=404, detail="Event not found")
        existing = (
            db.query(EventLikeRow)
            .filter_by(event_id=event_id, admin_email=admin["email"])
            .first()
        )
        if existing:
            db.delete(existing)
            liked = False
        else:
            db.add(EventLikeRow(event_id=event_id, admin_email=admin["email"]))
            liked = True
        db.commit()
        count = db.query(EventLikeRow).filter_by(event_id=event_id).count()
    return {"liked": liked, "like_count": count}


@api.get("/feed/{event_id}/comments")
def get_comments(event_id: int, admin: dict = Depends(current_admin)):
    with SessionLocal() as db:
        rows = (
            db.query(EventCommentRow)
            .filter_by(event_id=event_id)
            .order_by(EventCommentRow.id.asc())
            .all()
        )
    return [
        {
            "id": r.id,
            "admin_email": r.admin_email,
            "admin_name": r.admin_name or r.admin_email,
            "body": r.body,
            "created_at": _iso_utc(r.created_at),
        }
        for r in rows
    ]


@api.post("/feed/{event_id}/comments", status_code=201)
def add_comment(event_id: int, body: CommentIn, admin: dict = Depends(current_admin)):
    comment_body = body.body.strip()
    if not comment_body:
        raise HTTPException(status_code=400, detail="Comment cannot be empty")
    with SessionLocal() as db:
        if not db.get(FeedEventRow, event_id):
            raise HTTPException(status_code=404, detail="Event not found")
        row = EventCommentRow(
            event_id=event_id,
            admin_email=admin["email"],
            admin_name=admin.get("name"),
            body=comment_body,
        )
        db.add(row)
        db.commit()
        return {
            "id": row.id,
            "admin_email": row.admin_email,
            "admin_name": row.admin_name or row.admin_email,
            "body": row.body,
            "created_at": _iso_utc(row.created_at),
        }


# ---- Direct messages --------------------------------------------------------

@api.get("/messages", response_model=list[Message])
def get_messages(
    with_email: Optional[str] = Query(default=None),
    admin: dict = Depends(current_admin),
):
    """Messages involving you. With ?with_email=X, just the thread with X."""
    me_email = admin["email"]
    with SessionLocal() as db:
        q = db.query(MessageRow)
        if with_email:
            q = q.filter(
                ((MessageRow.sender == me_email) & (MessageRow.recipient == with_email))
                | ((MessageRow.sender == with_email) & (MessageRow.recipient == me_email))
            )
        else:
            q = q.filter(
                (MessageRow.sender == me_email) | (MessageRow.recipient == me_email)
            )
        rows = q.order_by(MessageRow.id.asc()).all()
    return [_serialize_message(r) for r in rows]


@api.post("/messages", response_model=Message, status_code=201)
def send_message(body: MessageIn, admin: dict = Depends(current_admin)):
    recipient = body.recipient.strip()
    text_body = body.body.strip()
    if not text_body:
        raise HTTPException(status_code=400, detail="Message cannot be empty")
    if recipient == admin["email"]:
        raise HTTPException(status_code=400, detail="You can't message yourself")
    if not is_active_admin(recipient):
        raise HTTPException(status_code=400, detail="Recipient must be an active admin")
    with SessionLocal() as db:
        row = MessageRow(sender=admin["email"], recipient=recipient, body=text_body)
        db.add(row)
        db.commit()
        return _serialize_message(row)


# ---- Red flags: SLA breaches computed live from portal tables --------------------
# No extra table needed — a red flag is a STATE (still pending past its SLA),
# read straight from pkdb. Add a rule by appending an entry here.

REDFLAG_WINDOW_DAYS = 45
COCO_GRN_ETA_BUFFER = 2   # red only once 2 days past the ETA


def _compute_coco_grn(conn, now: datetime) -> list[dict]:
    """CoCo GRN — PK's Pending-GRN-by-Location model. The LIST is every order
    still awaiting GRN (status approved/dispatched/invoiced/delivered,
    grn_completed_at NULL, empty composite parents stripped). An order becomes a
    RED FLAG only once it's 2+ days past its ETA (expected_delivery_date + 2d <
    now) — anchoring on the ETA, not the delivered-status, so vendors who never
    flip an order to 'delivered' are still caught (the ETA exists on approved
    orders too). Sub-order id, PO number and Sedna SO id carried for PK reconcile."""
    sql = (
        "SELECT io.id AS order_id, io.status AS state, "
        "io.expected_delivery_date AS eta, "
        "DATEDIFF(CURDATE(), io.expected_delivery_date) AS days_past_eta, "
        "(io.expected_delivery_date IS NOT NULL "
        " AND DATE_ADD(io.expected_delivery_date, INTERVAL :buf DAY) < NOW()) AS is_red, "
        "k.name AS kitchen, k.email AS kitchen_email, "
        "MAX(po.po_number) AS po_number, "
        "MAX(JSON_UNQUOTE(JSON_EXTRACT(po.sedna_tracking_data,'$.id'))) AS so_id, "
        "MAX(pv.name) AS vendor "
        "FROM pkdb.internal_orders io "
        "LEFT JOIN pkdb.coco_kitchens k ON k.id = io.kitchen_id "
        "LEFT JOIN pkdb.coco_purchase_orders po ON po.order_id = io.id "
        "LEFT JOIN pkdb.vendors pv ON pv.id = po.vendor_id "
        "WHERE io.status IN ('approved','dispatched','invoiced','delivered') "
        "AND io.grn_completed_at IS NULL "
        "AND NOT (EXISTS (SELECT 1 FROM pkdb.internal_orders c WHERE c.parent_order_id = io.id) "
        "         AND NOT EXISTS (SELECT 1 FROM pkdb.internal_order_items it WHERE it.order_id = io.id)) "
        "AND io.created_at >= :window "
        "GROUP BY io.id, io.status, io.expected_delivery_date, k.name, k.email "
        "ORDER BY is_red DESC, days_past_eta DESC, k.name ASC LIMIT 800"
    )
    try:
        rows = conn.execute(
            text(sql),
            {"buf": COCO_GRN_ETA_BUFFER, "window": now - timedelta(days=REDFLAG_WINDOW_DAYS)},
        ).mappings().all()
    except Exception:
        return []
    out = []
    for r in rows:
        red = bool(r["is_red"])
        dpe = r["days_past_eta"]
        out.append(
            {
                "entity": r["kitchen"] or "—",
                "contact_email": (r["kitchen_email"] or "").strip(),
                "po_number": r["po_number"],
                "order_id": r["order_id"],
                "so_id": r["so_id"],
                "vendor": (r["vendor"] or "").strip() or None,
                "eta": r["eta"],
                "days_overdue": max(0, int(dpe)) if (red and dpe is not None) else 0,
                "red": red,
                "state": r["state"],
                "ref": r["order_id"],
            }
        )
    return out


PARTNER_GRN_MAX_DAYS = 30   # ignore orders overdue more than a month — no point


def _compute_partner_grn(conn, now: datetime) -> list[dict]:
    """Partner orders completed but GRN not done beyond 2 days, due within the
    last month. Aligned with PartnerKart's Auto-GRN rule: when it's enabled we
    never flag orders completed before its effective_from (pre-rule legacy that
    Auto-GRN deliberately skips). Note Auto-GRN only clears fully 'completed'
    orders — 'partial_completed' ones never auto-clear and keep showing until
    someone GRNs them. The status is carried through as the type."""
    window = now - timedelta(days=PARTNER_GRN_MAX_DAYS)
    # Respect the Auto-GRN effective date so we match what it will actually
    # clear, instead of surfacing the pre-rule backlog it won't touch.
    try:
        s = conn.execute(
            text("SELECT enabled, effective_from FROM pkdb.partner_auto_grn_settings "
                 "ORDER BY id LIMIT 1")
        ).mappings().first()
        if s and s["enabled"] and s["effective_from"]:
            eff = s["effective_from"]
            eff_dt = datetime(eff.year, eff.month, eff.day) if hasattr(eff, "year") else None
            if eff_dt and eff_dt > window:
                window = eff_dt
    except Exception:
        pass
    sql = (
        "SELECT o.id AS order_id, o.parent_order_id AS so_ref, "
        "p.name AS partner, p.email AS partner_email, o.status AS state, "
        "o.updated_at AS since "
        "FROM pkdb.orders o LEFT JOIN pkdb.partners p ON p.id = o.partner_id "
        "WHERE o.status IN ('completed','partial_completed') "
        "AND o.grn_completed_at IS NULL "
        "AND o.updated_at < :cutoff AND o.updated_at >= :window "
        "ORDER BY o.updated_at DESC LIMIT 600"
    )
    try:
        rows = conn.execute(
            text(sql),
            {"cutoff": now - timedelta(days=2), "window": window},
        ).mappings().all()
    except Exception:
        return []
    out = []
    for r in rows:
        days = max(1, (now - r["since"]).days - 2) if r["since"] else 1
        out.append(
            {
                "entity": r["partner"] or "—",
                "contact_email": (r["partner_email"] or "").strip(),
                "po_number": None,
                "order_id": r["order_id"],
                "so_id": (f"#{r['so_ref']}" if r["so_ref"] else None),
                "vendor": None,
                "eta": None,
                "days_overdue": days,
                "state": r["state"],
                "ref": r["order_id"],
            }
        )
    return out


# Delhi upload paths are stored relative (e.g. /uploads/delhi/...); serve them
# off the PartnerKart host so the bill photo is a clickable link.
DELHI_UPLOAD_BASE = os.environ.get("DELHI_UPLOAD_BASE", "https://partner.kftpl.com")
DELHI_GRN_SLA_DAYS = 2   # an "ordered" batch older than this is GRN-overdue
DELHI_WINDOW_DAYS = 7


def _compute_delhi_grn(conn, now: datetime) -> list[dict]:
    """Delhi orders (grouped by order batch) with a problem in the last week:
    (a) GRN not done — still 'ordered' beyond the SLA, or (b) GRN done but no
    bill/invoice uploaded. Each flag carries its line items for drill-down and
    the uploaded bill photo URL when present."""
    sql = (
        "SELECT o.order_batch_id AS batch, o.id AS row_id, o.status, "
        "o.quantity, o.received_quantity, o.total_amount, o.unit_price, "
        "o.invoice_number, o.delivered_at, o.created_at, o.delivery_photo_url, "
        "k.name AS kitchen, k.email AS kitchen_email, v.name AS vendor, "
        "di.name AS item_name "
        "FROM pkdb.delhi_orders o "
        "LEFT JOIN pkdb.coco_kitchens k ON k.id = o.kitchen_id "
        "LEFT JOIN pkdb.delhi_vendors v ON v.id = o.vendor_id "
        "LEFT JOIN pkdb.delhi_order_items di ON di.id = o.item_id "
        "WHERE o.created_at >= :window "
        "ORDER BY o.order_batch_id, o.id"
    )
    try:
        rows = conn.execute(
            text(sql), {"window": now - timedelta(days=DELHI_WINDOW_DAYS)}
        ).mappings().all()
    except Exception:
        return []

    batches: dict = {}
    for r in rows:
        b = batches.setdefault(
            r["batch"],
            {
                "kitchen": r["kitchen"] or "—",
                "email": (r["kitchen_email"] or "").strip(),
                "vendor": (r["vendor"] or "").strip() or None,
                "created": r["created_at"],
                "delivered": r["delivered_at"],
                "invoice": None,
                "bill_url": None,
                "pending": 0,
                "amount": 0.0,
                "items": [],
                "ref_id": r["row_id"],
            },
        )
        if r["row_id"] < b["ref_id"]:
            b["ref_id"] = r["row_id"]
        if r["status"] == "ordered":
            b["pending"] += 1
        if r["invoice_number"] and str(r["invoice_number"]).strip():
            b["invoice"] = str(r["invoice_number"]).strip()
        if r["delivery_photo_url"] and not b["bill_url"]:
            b["bill_url"] = r["delivery_photo_url"]
        if r["delivered_at"] and (not b["delivered"] or r["delivered_at"] > b["delivered"]):
            b["delivered"] = r["delivered_at"]
        b["amount"] += float(r["total_amount"] or 0)
        b["items"].append(
            {
                "name": r["item_name"] or "—",
                "ordered": float(r["quantity"] or 0),
                "received": float(r["received_quantity"]) if r["received_quantity"] is not None else None,
                "amount": float(r["total_amount"] or 0),
            }
        )

    grn_cutoff = now - timedelta(days=DELHI_GRN_SLA_DAYS)
    out = []
    for bid, b in batches.items():
        if b["pending"] > 0:
            if not b["created"] or b["created"] > grn_cutoff:
                continue  # still within the receiving SLA — not a flag yet
            kind, state = "grn", "GRN pending"
            base = b["created"]
        elif not b["invoice"]:
            kind, state = "bill", "Bill pending"
            base = b["delivered"] or b["created"]
        else:
            continue  # received and billed — all good
        days = max(0, (now - base).days) if base else 0
        bill_url = None
        if b["bill_url"]:
            u = b["bill_url"]
            bill_url = u if u.startswith("http") else DELHI_UPLOAD_BASE + u
        out.append(
            {
                "entity": b["kitchen"],
                "contact_email": b["email"],
                "vendor": b["vendor"],
                "state": state,
                "kind": kind,
                "ident": f"Batch {str(bid)[:8]}",
                "amount": round(b["amount"], 2) or None,
                "bill_url": bill_url,
                "items": b["items"],
                "days_overdue": days,
                "po_number": None,
                "order_id": None,
                "so_id": None,
                "eta": None,
                "ref": b["ref_id"],  # a real delhi_orders row id in the batch
            }
        )
    out.sort(key=lambda f: (f["entity"].lower(), -f["days_overdue"]))
    return out


# ---- Kitchen launch (KLM / wodb1) --------------------------------------------
LAUNCH_OB_REMIND_DAYS = 10    # team reminds from day 10
LAUNCH_OB_DEADLINE_DAYS = 15  # onboarding deadline: 15 days from Aggregator OB


def _compute_launch_ob(conn, now: datetime) -> list[dict]:
    """Kitchen-launch onboarding SLA (KLM). Once 'Aggregator Onboarding' is
    completed, onboarding has a 15-day deadline (reminders from day 10). Flags
    launches still not live 10+ days on; RED once past the 15-day deadline.
    Grouped by launch manager so the laggards are obvious. Reads wodb1 — if the
    app's DB user lacks that grant the rule just returns empty (no crash)."""
    sql = (
        "SELECT agg.project_id AS ref, "
        "COALESCE(NULLIF(MAX(l.launch_manager),'None'),'Unassigned') AS manager, "
        "MAX(l.kac_name) AS kac_name, MAX(l.current_status) AS cur_status, "
        "MAX(agg.completed_at) AS ob_at, "
        "DATEDIFF(NOW(), MAX(agg.completed_at)) AS days_since, "
        "DATE_ADD(DATE(MAX(agg.completed_at)), INTERVAL :ddl DAY) AS deadline "
        "FROM wodb1.pk_milestone agg "
        "JOIN wodb1.pk_launch l ON l.id = agg.project_id "
        "WHERE agg.milestone_name IN ('Step 6: Aggregator Onboarding','OB request sent to aggregators') "
        "AND agg.status = 'completed' AND agg.completed_at >= (NOW() - INTERVAL 60 DAY) "
        "AND l.launch_completed_date IS NULL "
        "AND NOT EXISTS (SELECT 1 FROM wodb1.pk_milestone d WHERE d.project_id = agg.project_id "
        "  AND d.milestone_name IN ('Step 13: Take Kitchen Live','Launch') AND d.status = 'completed') "
        "GROUP BY agg.project_id "
        "HAVING days_since >= :remind "
        "ORDER BY days_since DESC LIMIT 400"
    )
    try:
        rows = conn.execute(
            text(sql),
            {"ddl": LAUNCH_OB_DEADLINE_DAYS, "remind": LAUNCH_OB_REMIND_DAYS},
        ).mappings().all()
    except Exception:
        return []
    out = []
    for r in rows:
        ds = int(r["days_since"] or 0)
        out.append(
            {
                "entity": r["manager"] or "Unassigned",
                "contact_email": "",
                "po_number": None,
                "order_id": None,
                "so_id": None,
                "ident": r["kac_name"] or f"Launch #{r['ref']}",
                "vendor": None,
                "eta": r["deadline"],  # the 15-day deadline
                "days_overdue": max(0, ds - LAUNCH_OB_DEADLINE_DAYS),
                "red": ds >= LAUNCH_OB_DEADLINE_DAYS,
                "state": r["cur_status"] or "onboarding",
                "ref": r["ref"],
            }
        )
    return out


LAUNCH_RM_CONFIRM_HRS = 24   # revenue team confirms date/time within 24h
LAUNCH_RM_CALL_HRS = 48      # call taken within 48h of the request


def _compute_launch_rm(conn, now: datetime) -> list[dict]:
    """RM & Revenue Call SLA (KLM). From when the RM & Revenue Call request/email
    is sent (pk_launch.rm_revenue_email), the revenue team confirms a date/time
    within 24h and the call is taken within 48h. Flags launches past 24h whose
    call isn't done; RED once past the 48h call window. NOTE: this stays empty
    until KLM starts stamping rm_revenue_email when the request goes out (it is
    not populated today — same family as onboarding_email_at)."""
    sql = (
        "SELECT l.id AS ref, "
        "COALESCE(NULLIF(l.launch_manager,'None'),'Unassigned') AS manager, "
        "l.kac_name, l.current_status, "
        "TIMESTAMPDIFF(HOUR, l.rm_revenue_email, NOW()) AS hrs, "
        "DATEDIFF(NOW(), l.rm_revenue_email) AS days_since, "
        "DATE_ADD(DATE(l.rm_revenue_email), INTERVAL 2 DAY) AS deadline, "
        "MAX(CASE WHEN m.milestone_name LIKE 'Step 14:%' THEN m.status END) AS rm_status, "
        "MAX(CASE WHEN m.milestone_name LIKE 'Step 14:%' "
        "  AND m.scheduled_datetime > '2000-01-01' THEN m.scheduled_datetime END) AS scheduled "
        "FROM wodb1.pk_launch l "
        "LEFT JOIN wodb1.pk_milestone m ON m.project_id = l.id "
        "WHERE l.rm_revenue_email IS NOT NULL "
        "AND l.rm_revenue_email >= (NOW() - INTERVAL 30 DAY) "
        "AND l.launch_completed_date IS NULL "
        "GROUP BY l.id, manager, l.kac_name, l.current_status, l.rm_revenue_email "
        "HAVING (rm_status IS NULL OR rm_status <> 'completed') "
        "AND TIMESTAMPDIFF(HOUR, l.rm_revenue_email, NOW()) >= :confirm "
        "ORDER BY hrs DESC LIMIT 400"
    )
    try:
        rows = conn.execute(text(sql), {"confirm": LAUNCH_RM_CONFIRM_HRS}).mappings().all()
    except Exception:
        return []
    out = []
    for r in rows:
        hrs = int(r["hrs"] or 0)
        confirmed = r["scheduled"] is not None
        red = hrs >= LAUNCH_RM_CALL_HRS
        state = "call not taken" if confirmed else "date/time not confirmed"
        out.append(
            {
                "entity": r["manager"] or "Unassigned",
                "contact_email": "",
                "po_number": None,
                "order_id": None,
                "so_id": None,
                "ident": r["kac_name"] or f"Launch #{r['ref']}",
                "vendor": None,
                "eta": r["deadline"],  # 48h call deadline
                "days_overdue": max(0, int(r["days_since"] or 0) - 2),
                "red": red,
                "state": state,
                "ref": r["ref"],
            }
        )
    return out


REDFLAG_RULES = [
    {
        "key": "coco_grn",
        "label": "CoCo orders — GRN pending (red = 2d+ past ETA)",
        "ref_label": "CoCo order",
        "party_label": "Kitchen",
        "group_by_kitchen": True,
        "count_red_only": True,
        "note": "PK Pending-GRN by kitchen. Red = 2+ days past ETA; rest awaiting/on-time.",
        "compute": _compute_coco_grn,
    },
    {
        "key": "partner_grn",
        "label": "Partner orders — GRN overdue (2 days from completion)",
        "ref_label": "Partner order",
        "party_label": "Partner",
        "group_by_kitchen": False,
        "note": "GRN pending >2d since Auto-GRN start. Partial orders never auto-clear.",
        "compute": _compute_partner_grn,
    },
    {
        "key": "delhi_grn",
        "label": "Delhi orders — GRN not done / bill not uploaded",
        "ref_label": "Delhi order",
        "party_label": "Kitchen",
        "group_by_kitchen": True,
        "note": "Last 7 days: GRN pending (>2d) or received but no bill uploaded",
        "compute": _compute_delhi_grn,
    },
    {
        "key": "launch_ob",
        "label": "Kitchen launch — onboarding overdue (15d from Aggregator OB)",
        "ref_label": "Launch",
        "party_label": "Launch manager",
        "group_by_kitchen": True,
        "count_red_only": True,
        "note": "Aggregator OB done, kitchen not live. Red = past the 15-day deadline; rest in the reminder window (from day 10). By launch manager.",
        "compute": _compute_launch_ob,
    },
    {
        "key": "launch_rm",
        "label": "Kitchen launch — RM & Revenue Call overdue (24h confirm / 48h call)",
        "ref_label": "Launch",
        "party_label": "Launch manager",
        "group_by_kitchen": True,
        "count_red_only": True,
        "note": "From RM & Revenue request: confirm date/time in 24h, take call in 48h. Red = past 48h. (Empty until KLM stamps rm_revenue_email.)",
        "compute": _compute_launch_rm,
    },
]

# Editable per-category reminder templates. {orders} {count} {due_date}
# {category} are substituted at send time.
DEFAULT_TEMPLATES = {
    "coco_grn": {
        "subject": "GRN pending — CoCo orders past ETA",
        "body": (
            "Team,\n\n"
            "The following CoCo orders are 2+ days past their delivery ETA with "
            "GRN still pending:\n\n{orders}\n\n"
            "Please complete the GRN by {due_date}.\n\n"
            "— Kouzina Live"
        ),
    },
    "partner_grn": {
        "subject": "Reminder: Please complete the GRN for your Kouzina order(s)",
        "body": (
            "Dear Partner,\n\n"
            "This is a reminder to complete the GRN (Goods Receipt Note) for the "
            "following order(s):\n\n{orders}\n\n"
            "Kindly complete the GRN on or before {due_date}. Any order not "
            "GRN-completed by this date will be treated as fully received and "
            "complete, and no grievances can be raised thereafter.\n\n"
            "Thank you,\nKouzina Team"
        ),
    },
    "delhi_grn": {
        "subject": "Delhi orders — GRN / bill pending",
        "body": (
            "Team,\n\n"
            "The following Delhi orders need attention — either the GRN is not "
            "done, or it is done but the bill has not been uploaded:\n\n{orders}\n\n"
            "Please complete the GRN and upload the bills by {due_date}.\n\n"
            "— Kouzina Live"
        ),
    },
    "launch_ob": {
        "subject": "Kitchen launch onboarding overdue",
        "body": (
            "Team,\n\n"
            "The following kitchen launches are past the onboarding deadline "
            "(15 days from Aggregator Onboarding) and are not yet live:\n\n{orders}\n\n"
            "Please push these to go-live by {due_date}.\n\n"
            "— Kouzina Live"
        ),
    },
    "launch_rm": {
        "subject": "RM & Revenue Call overdue",
        "body": (
            "Team,\n\n"
            "For the following launches the RM & Revenue Call is overdue "
            "(date/time to be confirmed within 24h, call taken within 48h):\n\n{orders}\n\n"
            "Please confirm and complete the call by {due_date}.\n\n"
            "— Kouzina Live"
        ),
    },
}

_redflag_cache: dict = {"at": None, "data": None}


def _rule_by_key(key: str) -> Optional[dict]:
    return next((r for r in REDFLAG_RULES if r["key"] == key), None)


def _flag_json(f: dict) -> dict:
    """Shape one breach for the API — carries every identifier the ops team
    needs to reconcile against PartnerKart (PO#, order#, Sedna SO#, vendor, ETA)."""
    eta = f.get("eta")
    eta_s = eta.isoformat() if hasattr(eta, "isoformat") else (eta or None)
    return {
        "ref": f["ref"],
        "entity": f["entity"],
        "state": f.get("state"),
        "po_number": f.get("po_number"),
        "order_id": f.get("order_id"),
        "so_id": f.get("so_id"),
        "vendor": f.get("vendor"),
        "eta": eta_s,
        "days_overdue": int(f.get("days_overdue") or 0),
        "red": f.get("red", True),  # False = shown for context, not a red flag
        # optional richer fields (Delhi orders): drill-down items, bill link, etc.
        "ident": f.get("ident"),
        "amount": f.get("amount"),
        "bill_url": f.get("bill_url"),
        "items": f.get("items"),
    }


def _subgroups_from_flags(flags: list[dict]) -> list[dict]:
    """Bucket breaches by entity (kitchen) so CoCo orders show under one kitchen
    heading — these are all COCO kitchens, so clubbing keeps the screen readable."""
    by: dict = {}
    for f in flags:
        key = f["entity"]
        g = by.setdefault(
            key,
            {"entity": key, "email": (f.get("contact_email") or "").strip(),
             "count": 0, "red": 0, "flags": []},
        )
        g["count"] += 1
        if f.get("red", True):
            g["red"] += 1
        g["flags"].append(_flag_json(f))
        if not g["email"] and f.get("contact_email"):
            g["email"] = f["contact_email"].strip()
    return sorted(by.values(), key=lambda g: (-g["red"], -g["count"], g["entity"]))


def _parties_from_flags(flags: list[dict]) -> list[dict]:
    """Distinct entities in a rule's breaches, with email + order count, so the
    UI can select which parties (kitchens/partners) to remind."""
    by = {}
    for f in flags:
        key = f["entity"]
        p = by.setdefault(
            key, {"entity": key, "email": (f.get("contact_email") or "").strip(), "count": 0}
        )
        p["count"] += 1
        if not p["email"] and f.get("contact_email"):
            p["email"] = f["contact_email"].strip()
    return sorted(by.values(), key=lambda p: (-p["count"], p["entity"]))


def _template_for(rule_key: str) -> dict:
    with SessionLocal() as db:
        row = db.get(RedflagTemplateRow, rule_key)
    base = DEFAULT_TEMPLATES.get(rule_key, {"subject": "", "body": "{orders}"})
    if row and (row.subject or row.body):
        return {"subject": row.subject or base["subject"], "body": row.body or base["body"],
                "is_custom": True}
    return {**base, "is_custom": False}


def _render(tmpl: str, ctx: dict) -> str:
    out = tmpl
    for k, v in ctx.items():
        out = out.replace("{" + k + "}", str(v))
    return out


@api.get("/redflags")
def redflags(admin: dict = Depends(current_admin)):
    now = datetime.now(timezone.utc).replace(tzinfo=None)
    if (
        _redflag_cache["at"]
        and (now - _redflag_cache["at"]).total_seconds() < 60
        and _redflag_cache["data"]
    ):
        return _redflag_cache["data"]

    def _rule_out(rule: dict, flags: list[dict]) -> dict:
        # Some rules (CoCo) show a full pending LIST but only the red subset is
        # a flag; the count + reminders track the red ones.
        red_only = rule.get("count_red_only")
        reds = [f for f in flags if f.get("red", True)]
        count = len(reds) if red_only else len(flags)
        out = {
            "key": rule["key"],
            "label": rule["label"],
            "ref_label": rule["ref_label"],
            "party_label": rule.get("party_label", "Party"),
            "note": rule.get("note", ""),
            "window_days": REDFLAG_WINDOW_DAYS,
            "grouped": bool(rule.get("group_by_kitchen")),
            "count": count,
            "pending": len(flags),
            "red_only": bool(red_only),
            "parties": _parties_from_flags(reds if red_only else flags),
            "flags": [_flag_json(f) for f in flags],
        }
        if rule.get("group_by_kitchen"):
            out["subgroups"] = _subgroups_from_flags(flags)
        return out

    rules_out, total = [], 0
    if IS_MYSQL:
        with engine.connect() as conn:
            for rule in REDFLAG_RULES:
                flags = rule["compute"](conn, now)
                ro = _rule_out(rule, flags)
                total += ro["count"]
                rules_out.append(ro)
    else:  # local dev demo
        coco_demo = [
            {"entity": "E-CITY", "contact_email": "ecity@kftpl.com", "po_number": "COPO/26-27/0142",
             "order_id": 1988, "so_id": "SO26-AAJCK-001188", "vendor": "Sedna Retail",
             "eta": (now - timedelta(days=4)).date(), "days_overdue": 4, "red": True,
             "state": "invoiced", "ref": 1988},
            {"entity": "E-CITY", "contact_email": "ecity@kftpl.com", "po_number": "COPO/26-27/0151",
             "order_id": 2001, "so_id": None, "vendor": "Local Traders",
             "eta": (now + timedelta(days=1)).date(), "days_overdue": 0, "red": False,
             "state": "approved", "ref": 2001},
            {"entity": "KLP HSR", "contact_email": "klphsr@kftpl.com", "po_number": "COPO/26-27/0133",
             "order_id": 1777, "so_id": "SO26-AAJCK-001177", "vendor": "Sedna Retail",
             "eta": (now - timedelta(days=6)).date(), "days_overdue": 6, "red": True,
             "state": "dispatched", "ref": 1777},
        ]
        partner_demo = [
            {"entity": "HUNMONI DUTTA", "contact_email": "hunmoni@example.com", "po_number": None,
             "order_id": 37474, "so_id": "#37400", "vendor": None, "eta": None,
             "days_overdue": 2, "state": "completed", "ref": 37474},
        ]
        delhi_demo = [
            {"entity": "DLF", "contact_email": "dlf@kftpl.com", "vendor": "Sai Enterprises",
             "state": "GRN pending", "kind": "grn", "ident": "Batch cf5f5e5a", "amount": 4200.0,
             "bill_url": None, "days_overdue": 3, "ref": 90101,
             "items": [{"name": "Coal Supply (DLF)", "ordered": 5, "received": None, "amount": 4200}]},
            {"entity": "DLF", "contact_email": "dlf@kftpl.com", "vendor": "Madhav Gas",
             "state": "Bill pending", "kind": "bill", "ident": "Batch 9ac81137", "amount": 1800.0,
             "bill_url": "https://partner.kftpl.com/uploads/delhi/delhi_2209_1783462777.jpg",
             "days_overdue": 2, "ref": 90102,
             "items": [{"name": "GAS CYLINDER (19 KGS)", "ordered": 2, "received": 2, "amount": 1800}]},
        ]
        launch_demo = [
            {"entity": "Satyam", "contact_email": "", "ident": "RF-UP-LKO-ALIGANJ0-1",
             "eta": (now - timedelta(days=5)).date(), "days_overdue": 5, "red": True,
             "state": "WIP", "ref": 415},
            {"entity": "Satyam", "contact_email": "", "ident": "RX-KA-BNG-RICHESGA-1",
             "eta": (now + timedelta(days=2)).date(), "days_overdue": 0, "red": False,
             "state": "onboarding", "ref": 421},
        ]
        rules_out = [
            _rule_out(REDFLAG_RULES[0], coco_demo),
            _rule_out(REDFLAG_RULES[1], partner_demo),
            _rule_out(REDFLAG_RULES[2], delhi_demo),
            _rule_out(REDFLAG_RULES[3], launch_demo),
        ]
        total = sum(r["count"] for r in rules_out)
    data = {
        "total": total,
        "rules": rules_out,
        "generated_at": _iso_utc(now),
        "email_enabled": email_enabled(),
    }
    _redflag_cache["at"] = now
    _redflag_cache["data"] = data
    return data


@api.get("/redflags/templates")
def get_templates(admin: dict = Depends(current_admin)):
    return [
        {"rule_key": rule["key"], "label": rule["label"], **_template_for(rule["key"])}
        for rule in REDFLAG_RULES
    ]


class TemplateIn(BaseModel):
    subject: str = Field(..., max_length=300)
    body: str = Field(..., max_length=6000)


@api.put("/redflags/templates/{rule_key}")
def save_template(rule_key: str, payload: TemplateIn, admin: dict = Depends(current_admin)):
    if not _rule_by_key(rule_key):
        raise HTTPException(status_code=404, detail="Unknown category")
    with SessionLocal() as db:
        row = db.get(RedflagTemplateRow, rule_key)
        if not row:
            row = RedflagTemplateRow(rule_key=rule_key)
            db.add(row)
        row.subject = payload.subject.strip()
        row.body = payload.body.strip()
        row.updated_by = admin["email"]
        row.updated_at = datetime.now(timezone.utc).replace(tzinfo=None)
        db.commit()
    return {"ok": True}


# ---- Flag a single breach to admins (in-app + email) ----------------------------

class FlagIn(BaseModel):
    ref: int
    entity: str = Field(..., max_length=255)
    label: str = Field(..., max_length=64)
    state: Optional[str] = None
    days_overdue: int = 0
    recipients: list[str] = Field(..., min_length=1)
    note: Optional[str] = Field(default=None, max_length=1000)


@api.post("/redflags/flag")
def flag_to_admins(
    payload: FlagIn, background_tasks: BackgroundTasks, admin: dict = Depends(current_admin)
):
    recipients = [e for e in dict.fromkeys(payload.recipients) if is_active_admin(e)]
    if not recipients:
        raise HTTPException(status_code=400, detail="Pick at least one admin")
    who = admin.get("name") or admin["email"]
    what = (payload.state or "").strip() or "GRN overdue"
    line = (
        f"🚩 {payload.label} #{payload.ref} · {payload.entity} · "
        f"{payload.days_overdue}d · {what}"
    )
    if payload.note:
        line += f" — {payload.note.strip()}"
    with SessionLocal() as db:
        for r in recipients:
            db.add(MessageRow(sender=admin["email"], recipient=r, body=line, is_private=0))
        db.commit()
    emit_live_event(
        who, "redflag_raised",
        f"flagged {payload.entity} · {payload.label} #{payload.ref}",
        {"module": "Red Flags"},
    )
    emailed = False
    if email_enabled():
        subject = f"[Kouzina Live] Red flag: {payload.entity} — {payload.label} #{payload.ref}"
        body = f"{line}\n\nFlagged by {who} via Kouzina Live.\nOpen: https://live.kftpl.com"
        background_tasks.add_task(send_email, recipients, subject, body)
        emailed = True
    return {"ok": True, "messaged": len(recipients), "emailed": emailed}


# ---- Send a reminder for a whole category (editable template) -------------------

class GroupSendIn(BaseModel):
    rule_key: str
    subject: str = Field(..., max_length=300)
    body: str = Field(..., max_length=6000)
    due_date: Optional[str] = None       # YYYY-MM-DD, fills {due_date}
    recipients: list[str] = []           # admins / any emails for one combined summary
    party_emails: list[str] = []         # specific kitchens/partners to mail individually
    save_template: bool = True           # persist the edited template for the category


def _orders_block(flags: list[dict], ref_label: str) -> str:
    lines = []
    for f in flags:
        ids = []
        if f.get("order_id"):
            ids.append(f"Order #{f['order_id']}")
        if f.get("po_number"):
            ids.append(f"PO {f['po_number']}")
        if f.get("so_id"):
            ids.append(f"SO {f['so_id']}")
        if f.get("ident"):
            ids.append(f["ident"])
        id_s = " · ".join(ids) or f"{ref_label} #{f.get('ref')}"
        parts = [f"• {f['entity']} — {id_s}"]
        if f.get("vendor"):
            parts.append(f"[{f['vendor']}]")
        if f.get("state"):
            parts.append(f"— {f['state']}")
        eta = f.get("eta")
        if eta:
            eta_s = eta.strftime("%d %b %Y") if hasattr(eta, "strftime") else str(eta)
            parts.append(f"— ETA {eta_s}")
        parts.append(f"— {f.get('days_overdue', 0)}d")
        if f.get("bill_url"):
            parts.append(f"— bill: {f['bill_url']}")
        lines.append(" ".join(parts))
    return "\n".join(lines) if lines else "(none)"


@api.post("/redflags/send-group")
def send_group_reminder(
    payload: GroupSendIn,
    background_tasks: BackgroundTasks,
    admin: dict = Depends(current_admin),
):
    rule = _rule_by_key(payload.rule_key)
    if not rule:
        raise HTTPException(status_code=404, detail="Unknown category")
    if payload.save_template:
        with SessionLocal() as db:
            row = db.get(RedflagTemplateRow, payload.rule_key) or RedflagTemplateRow(
                rule_key=payload.rule_key
            )
            row.subject = payload.subject.strip()
            row.body = payload.body.strip()
            row.updated_by = admin["email"]
            row.updated_at = datetime.now(timezone.utc).replace(tzinfo=None)
            db.merge(row)
            db.commit()

    now = datetime.now(timezone.utc).replace(tzinfo=None)
    due = payload.due_date
    if due:
        try:
            due = datetime.fromisoformat(due).strftime("%d %b %Y")
        except ValueError:
            due = payload.due_date
    else:
        due = (now + timedelta(days=2)).strftime("%d %b %Y")

    flags = []
    if IS_MYSQL:
        with engine.connect() as conn:
            flags = rule["compute"](conn, now)
    # Red-only rules (CoCo) remind about the past-ETA orders, not the whole
    # pending list that's shown for context.
    if rule.get("count_red_only"):
        flags = [f for f in flags if f.get("red", True)]
    who = admin.get("name") or admin["email"]

    def ctx_for(subset):
        return {
            "orders": _orders_block(subset, rule["ref_label"]),
            "count": len(subset),
            "category": rule["label"],
            "due_date": due,
        }

    sent, messaged, party_mails = 0, 0, 0

    # 1) One combined summary email to chosen recipients (admins / ops / custom).
    recips = [e.strip() for e in dict.fromkeys(payload.recipients) if e and "@" in e]
    if recips and email_enabled():
        background_tasks.add_task(
            send_email,
            recips,
            _render(payload.subject, ctx_for(flags)),
            _render(payload.body, ctx_for(flags)),
        )
        sent += len(recips)
    admin_recips = [e for e in recips if is_active_admin(e)]
    if admin_recips:
        note = f"🚩 {rule['label']} — {len(flags)} pending. Reminder sent."
        with SessionLocal() as db:
            for e in admin_recips:
                db.add(MessageRow(sender=admin["email"], recipient=e, body=note, is_private=0))
            db.commit()
            messaged = len(admin_recips)

    # 2) Email SPECIFIC parties (kitchens/partners) their OWN orders individually.
    wanted = {e.strip().lower() for e in payload.party_emails if e and "@" in e}
    if wanted and email_enabled():
        by_email: dict[str, list] = {}
        for f in flags:
            em = (f.get("contact_email") or "").strip()
            if em and em.lower() in wanted:
                by_email.setdefault(em, []).append(f)
        for em, subset in by_email.items():
            background_tasks.add_task(
                send_email,
                [em],
                _render(payload.subject, ctx_for(subset)),
                _render(payload.body, ctx_for(subset)),
            )
            party_mails += 1
        sent += party_mails

    emit_live_event(
        who, "redflag_reminder_sent",
        f"sent {rule['ref_label']} GRN reminder ({len(flags)} pending)",
        {"module": "Red Flags"},
    )
    return {
        "ok": True,
        "pending": len(flags),
        "emailed": sent,
        "messaged": messaged,
        "party_emails": party_mails,
        "email_enabled": email_enabled(),
    }


# ---- Programs -------------------------------------------------------------------

class ProgramIn(BaseModel):
    name: str = Field(..., min_length=2, max_length=255)
    objective: Optional[str] = Field(default=None, max_length=2000)
    description: Optional[str] = Field(default=None, max_length=5000)
    owner_email: Optional[str] = None
    department: Optional[str] = Field(default=None, max_length=255)
    eta: Optional[str] = None  # YYYY-MM-DD


class ProgramPatch(BaseModel):
    name: Optional[str] = Field(default=None, min_length=2, max_length=255)
    objective: Optional[str] = Field(default=None, max_length=2000)
    description: Optional[str] = Field(default=None, max_length=5000)
    owner_email: Optional[str] = None
    department: Optional[str] = Field(default=None, max_length=255)
    eta: Optional[str] = None
    status: Optional[str] = None
    active: Optional[bool] = None


def _parse_eta(eta: Optional[str]):
    if not eta:
        return None
    try:
        return datetime.fromisoformat(eta)
    except ValueError:
        raise HTTPException(status_code=400, detail="ETA must be YYYY-MM-DD")


def _owner_fields(owner_email: Optional[str]) -> tuple[Optional[str], Optional[str]]:
    if not owner_email:
        return None, None
    names = resolve_names({owner_email})
    return owner_email, names.get(owner_email) or owner_email.split("@")[0]


def _serialize_program(
    p: ProgramRow, updates_count: int = 0, last_update: Optional[dict] = None
) -> dict:
    return {
        "id": p.id,
        "name": p.name,
        "objective": p.objective,
        "description": p.description,
        "owner_email": p.owner_email,
        "owner_name": p.owner_name,
        "department": p.department,
        "eta": p.eta.date().isoformat() if p.eta else None,
        "status": p.status,
        "active": bool(p.active),
        "updates_count": updates_count,
        "last_update": last_update,
        "created_at": _iso_utc(p.created_at),
    }


@api.get("/programs")
def list_programs(
    limit: int = Query(default=10, le=50),
    offset: int = Query(default=0, ge=0),
    department: Optional[str] = Query(default=None),
    admin: dict = Depends(current_admin),
):
    with SessionLocal() as db:
        q = db.query(ProgramRow)
        if department:
            q = q.filter(ProgramRow.department == department)
        total = q.count()
        rows = (
            q.order_by(ProgramRow.active.desc(), ProgramRow.eta.asc(), ProgramRow.id.asc())
            .offset(offset)
            .limit(limit)
            .all()
        )
        ids = [p.id for p in rows]
        counts = {}
        latest = {}
        if ids:
            counts = dict(
                db.query(ProgramUpdateRow.program_id, func.count())
                .filter(ProgramUpdateRow.program_id.in_(ids))
                .group_by(ProgramUpdateRow.program_id)
            )
            # newest update per program (id is monotonic, so max(id) = latest)
            newest_ids = [
                r[0]
                for r in db.query(func.max(ProgramUpdateRow.id))
                .filter(ProgramUpdateRow.program_id.in_(ids))
                .group_by(ProgramUpdateRow.program_id)
                .all()
            ]
            for u in db.query(ProgramUpdateRow).filter(ProgramUpdateRow.id.in_(newest_ids)):
                latest[u.program_id] = {
                    "author_name": u.author_name or u.author_email,
                    "body": u.body,
                    "created_at": _iso_utc(u.created_at),
                }
    return {
        "programs": [
            _serialize_program(p, counts.get(p.id, 0), latest.get(p.id)) for p in rows
        ],
        "total": total,
    }


@api.post("/programs", status_code=201)
def create_program(payload: ProgramIn, admin: dict = Depends(current_admin)):
    owner_email, owner_name = _owner_fields(payload.owner_email)
    with SessionLocal() as db:
        # Department follows the owner unless set explicitly (department page).
        dept = (payload.department or "").strip() or None
        if not dept and owner_email:
            p = (
                db.query(UserProfileRow)
                .filter(func.lower(UserProfileRow.email) == owner_email.lower())
                .first()
            )
            dept = p.department if p else None
        row = ProgramRow(
            name=payload.name.strip(),
            objective=(payload.objective or "").strip() or None,
            description=(payload.description or "").strip() or None,
            owner_email=owner_email,
            owner_name=owner_name,
            department=dept,
            eta=_parse_eta(payload.eta),
            status="not_started",
            created_by=admin["email"],
        )
        db.add(row)
        db.commit()
        result = _serialize_program(row)
    emit_live_event(
        admin.get("name") or admin["email"],
        "program_created",
        "created program",
        {"name": result["name"], "module": "Programs", "eta": result["eta"], "owner": owner_name},
    )
    return result


@api.patch("/programs/{program_id}")
def patch_program(program_id: int, payload: ProgramPatch, admin: dict = Depends(current_admin)):
    actor = admin.get("name") or admin["email"]
    with SessionLocal() as db:
        row = db.get(ProgramRow, program_id)
        if not row:
            raise HTTPException(status_code=404, detail="Program not found")
        events = []
        if payload.status is not None and payload.status != row.status:
            if payload.status not in STATUS_LABELS:
                raise HTTPException(status_code=400, detail="Bad status")
            events.append(
                (
                    "program_status_changed",
                    "changed program status",
                    {
                        "name": row.name,
                        "module": "Programs",
                        "status": {
                            "from": STATUS_LABELS[row.status],
                            "to": STATUS_LABELS[payload.status],
                        },
                    },
                )
            )
            row.status = payload.status
        if payload.active is not None and bool(row.active) != payload.active:
            row.active = 1 if payload.active else 0
            verb = "reactivated program" if payload.active else "deactivated program"
            events.append(
                (
                    "program_deactivated" if not payload.active else "program_reactivated",
                    verb,
                    {"name": row.name, "module": "Programs"},
                )
            )
        edited = False
        if payload.name is not None and payload.name.strip() != row.name:
            row.name = payload.name.strip()
            edited = True
        if payload.objective is not None and payload.objective.strip() != (row.objective or ""):
            row.objective = payload.objective.strip() or None
            edited = True
        if payload.description is not None and payload.description.strip() != (row.description or ""):
            row.description = payload.description.strip() or None
            edited = True
        if payload.owner_email is not None:
            row.owner_email, row.owner_name = _owner_fields(payload.owner_email or None)
            edited = True
            if payload.department is None:  # department follows the owner
                p = (
                    db.query(UserProfileRow)
                    .filter(func.lower(UserProfileRow.email) == (row.owner_email or "").lower())
                    .first()
                ) if row.owner_email else None
                row.department = p.department if p else None
        if payload.department is not None:
            row.department = payload.department.strip() or None
            edited = True
        if payload.eta is not None:
            row.eta = _parse_eta(payload.eta)
            edited = True
        if edited:
            events.append(
                (
                    "program_edited",
                    "edited program",
                    {"name": row.name, "module": "Programs", "owner": row.owner_name,
                     "eta": row.eta.date().isoformat() if row.eta else None},
                )
            )
        db.commit()
        result = _serialize_program(row)
    for action, summary, details in events:
        emit_live_event(actor, action, summary, details)
    return result


@api.get("/programs/{program_id}/updates")
def program_updates(program_id: int, admin: dict = Depends(current_admin)):
    with SessionLocal() as db:
        rows = (
            db.query(ProgramUpdateRow)
            .filter_by(program_id=program_id)
            .order_by(ProgramUpdateRow.id.desc())
            .limit(50)
            .all()
        )
    return [
        {
            "id": r.id,
            "author_name": r.author_name or r.author_email,
            "body": r.body,
            "created_at": _iso_utc(r.created_at),
        }
        for r in rows
    ]


@api.post("/programs/{program_id}/updates", status_code=201)
def add_program_update(
    program_id: int, payload: CommentIn, admin: dict = Depends(current_admin)
):
    body = payload.body.strip()
    if not body:
        raise HTTPException(status_code=400, detail="Update cannot be empty")
    actor = admin.get("name") or admin["email"]
    with SessionLocal() as db:
        prog = db.get(ProgramRow, program_id)
        if not prog:
            raise HTTPException(status_code=404, detail="Program not found")
        db.add(
            ProgramUpdateRow(
                program_id=program_id,
                author_email=admin["email"],
                author_name=admin.get("name"),
                body=body,
            )
        )
        db.commit()
        prog_name = prog.name
    emit_live_event(
        actor,
        "program_update_posted",
        f"posted update · “{body[:70]}”",
        {"name": prog_name, "module": "Programs"},
    )
    return {"ok": True}


# ---- Anonymous feedback -----------------------------------------------------------

class ActionIn(BaseModel):
    action_item: str = Field(..., min_length=1, max_length=1000)


@api.get("/feedback")
def list_feedback(admin: dict = Depends(current_admin)):
    with SessionLocal() as db:
        rows = db.query(FeedbackRow).order_by(FeedbackRow.id.desc()).limit(100).all()
    return [
        {
            "id": r.id,
            "body": r.body,
            "status": r.status,
            "action_item": r.action_item,
            "action_by": r.action_by,
            "action_at": _iso_utc(r.action_at),
            "created_at": _iso_utc(r.created_at),
        }
        for r in rows
    ]


@api.post("/feedback", status_code=201)
def add_feedback(payload: CommentIn, admin: dict = Depends(current_admin)):
    body = payload.body.strip()
    if not body:
        raise HTTPException(status_code=400, detail="Feedback cannot be empty")
    # The author is intentionally NOT stored anywhere.
    with SessionLocal() as db:
        db.add(FeedbackRow(body=body))
        db.commit()
    emit_live_event(
        "Anonymous",
        "feedback_received",
        f"shared feedback · “{body[:70]}”",
        {"module": "Feedback"},
    )
    return {"ok": True}


@api.post("/feedback/{feedback_id}/action")
def set_feedback_action(
    feedback_id: int, payload: ActionIn, admin: dict = Depends(current_admin)
):
    actor = admin.get("name") or admin["email"]
    item = payload.action_item.strip()
    with SessionLocal() as db:
        row = db.get(FeedbackRow, feedback_id)
        if not row:
            raise HTTPException(status_code=404, detail="Feedback not found")
        row.action_item = item
        row.action_by = actor
        row.action_at = datetime.now(timezone.utc).replace(tzinfo=None)
        row.status = "actioned"
        db.commit()
    emit_live_event(
        actor,
        "feedback_action_added",
        f"added action item · “{item[:70]}”",
        {"module": "Feedback"},
    )
    return {"ok": True}


app.mount("/api", api)

# ---- Static frontend (production) ------------------------------------------
FRONTEND_DIR = Path(os.environ.get("FRONTEND_DIR", Path(__file__).parent / "static"))
if FRONTEND_DIR.is_dir():
    app.mount("/", StaticFiles(directory=FRONTEND_DIR, html=True), name="frontend")
